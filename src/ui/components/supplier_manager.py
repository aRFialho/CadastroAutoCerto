"""Interface de gerenciamento de fornecedores"""

import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from pathlib import Path
from typing import Optional

from ...core.supplier_database import SupplierDatabase, Supplier
from ...utils.logger import get_logger

logger = get_logger("supplier_manager")


class AccessModeDialog:
    """Diálogo para escolher modo de acesso"""

    def __init__(self, parent):
        self.parent = parent
        self.result = None
        self.setup_dialog()

    def setup_dialog(self):
        """Configura o diálogo"""
        self.dialog = ctk.CTkToplevel(self.parent)
        self.dialog.title("🔐 Modo de Acesso")
        self.dialog.geometry("400x250")
        self.dialog.minsize(350, 200)

        # Centralizar na tela
        self.dialog.transient(self.parent)
        self.dialog.grab_set()

        # Impedir redimensionamento
        self.dialog.resizable(False, False)

        # Título
        title_label = ctk.CTkLabel(
            self.dialog,
            text="🗄️ Banco de Fornecedores",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title_label.pack(pady=(20, 10))

        # Subtítulo
        subtitle_label = ctk.CTkLabel(
            self.dialog,
            text="Escolha o modo de acesso:",
            font=ctk.CTkFont(size=14),
            text_color=("gray60", "gray40")
        )
        subtitle_label.pack(pady=(0, 20))

        # Frame dos botões
        buttons_frame = ctk.CTkFrame(self.dialog, fg_color="transparent")
        buttons_frame.pack(expand=True, fill="both", padx=40, pady=20)

        # Botão Admin
        admin_btn = ctk.CTkButton(
            buttons_frame,
            text="🔧 Abrir como Admin",
            command=self.select_admin,
            font=ctk.CTkFont(size=16, weight="bold"),
            height=50,
            corner_radius=10
        )
        admin_btn.pack(fill="x", pady=(0, 15))

        # Descrição Admin
        admin_desc = ctk.CTkLabel(
            buttons_frame,
            text="• Adicionar, editar e excluir fornecedores\n• Importar e exportar dados\n• Requer senha de administrador",
            font=ctk.CTkFont(size=12),
            text_color=("gray60", "gray40"),
            justify="left"
        )
        admin_desc.pack(pady=(0, 20))

        # Botão Busca
        search_btn = ctk.CTkButton(
            buttons_frame,
            text="👁️ Apenas Busca",
            command=self.select_search,
            font=ctk.CTkFont(size=16),
            height=50,
            corner_radius=10,
            fg_color=("gray60", "gray40"),
            hover_color=("gray50", "gray30")
        )
        search_btn.pack(fill="x", pady=(0, 15))

        # Descrição Busca
        search_desc = ctk.CTkLabel(
            buttons_frame,
            text="• Visualizar lista de fornecedores\n• Buscar e filtrar dados\n• Somente leitura",
            font=ctk.CTkFont(size=12),
            text_color=("gray60", "gray40"),
            justify="left"
        )
        search_desc.pack()

        # Centralizar janela
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")

    def select_admin(self):
        """Seleciona modo administrador"""
        self.result = "admin"
        self.dialog.destroy()

    def select_search(self):
        """Seleciona modo busca"""
        self.result = "search"
        self.dialog.destroy()


class SupplierManagerWindow:
    """Janela de gerenciamento de fornecedores"""

    ADMIN_PASSWORD = "172839"
    def __init__(self, parent, db_path: Path):
        self.parent = parent
        self.db_path = Path(db_path)
        self.db = None
        self.db_available = False
        self.selected_supplier = None

        # ✅ Modo robusto: se não existir arquivo / não der pra abrir, não quebra o app
        if not self.db_path.exists():
            messagebox.showwarning(
                "Fornecedores indisponível",
                f"Banco de fornecedores não encontrado:\\n{self.db_path}\\n\\n"
                "Este módulo ficará indisponível até o banco ser criado/baixado."
            )
            logger.warning(f"Banco de fornecedores ausente: {self.db_path}")
            return

        try:
            self.db = SupplierDatabase(self.db_path)
            self.db_available = True
        except Exception as e:
            messagebox.showwarning(
                "Fornecedores indisponível",
                f"Não foi possível abrir o banco de fornecedores:\\n{self.db_path}\\n\\nErro: {e}"
            )
            logger.error(f"Erro ao abrir banco de fornecedores: {e}")
            self.db = None
            self.db_available = False
            return

    def setup_window(self):
        """Configura a janela"""
        # Título baseado no modo
        if self.admin_mode:
            title = "🔧 Gerenciamento de Fornecedores (Admin)"
            window_title = "🗄️ Fornecedores - Modo Administrador"
        else:
            title = "👁️ Consulta de Fornecedores (Somente Leitura)"
            window_title = "🗄️ Fornecedores - Modo Consulta"

        self.window = ctk.CTkToplevel(self.parent)
        self.window.title(window_title)
        self.window.geometry("1400x800")  # ✅ AUMENTAR LARGURA PARA NOVA COLUNA
        self.window.minsize(1200, 700)

        # Centralizar na tela
        self.window.transient(self.parent)
        self.window.grab_set()

        self.create_header(title)
        self.create_toolbar()
        self.create_main_content()
        self.create_footer()

    def create_header(self, title: str = None):
        """Cria cabeçalho"""
        header_frame = ctk.CTkFrame(self.window, height=80)
        header_frame.pack(fill="x", padx=20, pady=(20, 10))
        header_frame.pack_propagate(False)

        if not title:
            title = "🗄️ Banco de Dados de Fornecedores"

        title_label = ctk.CTkLabel(
            header_frame,
            text=title,
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(expand=True)

        # ✅ INDICADOR DE MODO
        if not self.admin_mode:
            mode_label = ctk.CTkLabel(
                header_frame,
                text="👁️ MODO SOMENTE LEITURA",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=("orange", "orange")
            )
            mode_label.pack(pady=(0, 5))

    def create_toolbar(self):
        """Cria barra de ferramentas"""
        toolbar_frame = ctk.CTkFrame(self.window)
        toolbar_frame.pack(fill="x", padx=20, pady=(0, 10))

        # Botões principais
        buttons_frame = ctk.CTkFrame(toolbar_frame, fg_color="transparent")
        buttons_frame.pack(side="left", fill="x", expand=True, padx=15, pady=15)

        # ✅ BOTÕES CONDICIONAIS BASEADOS NO MODO
        if self.admin_mode:
            self.add_btn = ctk.CTkButton(
                buttons_frame,
                text="➕ Adicionar",
                command=self.add_supplier,
                width=100,
                height=35
            )
            self.add_btn.pack(side="left", padx=(0, 10))

            self.edit_btn = ctk.CTkButton(
                buttons_frame,
                text="✏️ Editar",
                command=self.edit_supplier,
                width=100,
                height=35,
                state="disabled"
            )
            self.edit_btn.pack(side="left", padx=(0, 10))

            self.delete_btn = ctk.CTkButton(
                buttons_frame,
                text="🗑️ Excluir",
                command=self.delete_supplier,
                width=100,
                height=35,
                state="disabled"
            )
            self.delete_btn.pack(side="left", padx=(0, 10))

            self.import_btn = ctk.CTkButton(
                buttons_frame,
                text="📥 Importar",
                command=self.import_suppliers,
                width=100,
                height=35
            )
            self.import_btn.pack(side="left", padx=(0, 10))

            self.example_btn = ctk.CTkButton(
                buttons_frame,
                text="Exemplo",
                command=self.create_example_file,
                width=100,
                height=35
            )
            self.example_btn.pack(side="left", padx=(0, 10))
        else:
            # ✅ MODO SOMENTE LEITURA - SEM BOTÕES DE EDIÇÃO
            info_label = ctk.CTkLabel(
                buttons_frame,
                text="👁️ Modo somente leitura - Use busca para encontrar fornecedores",
                font=ctk.CTkFont(size=14),
                text_color=("gray60", "gray40")
            )
            info_label.pack(side="left", padx=(0, 20))

        # ✅ BOTÃO EXPORTAR SEMPRE DISPONÍVEL
        self.export_btn = ctk.CTkButton(
            buttons_frame,
            text="📤 Exportar",
            command=self.export_suppliers,
            width=100,
            height=35
        )
        self.export_btn.pack(side="left", padx=(0, 10))

        # Campo de busca (sempre disponível)
        search_frame = ctk.CTkFrame(toolbar_frame, fg_color="transparent")
        search_frame.pack(side="right", padx=15, pady=15)

        ctk.CTkLabel(search_frame, text="🔍 Buscar:").pack(side="left", padx=(0, 5))

        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)

        self.search_entry = ctk.CTkEntry(
            search_frame,
            textvariable=self.search_var,
            placeholder_text="Nome do fornecedor...",
            width=200
        )
        self.search_entry.pack(side="left", padx=(0, 5))

        ctk.CTkButton(
            search_frame,
            text="🔄",
            command=self.load_suppliers,
            width=30,
            height=30
        ).pack(side="left")

    def create_main_content(self):
        """Cria conteúdo principal"""
        main_frame = ctk.CTkFrame(self.window)
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        # Treeview para lista de fornecedores
        tree_frame = ctk.CTkFrame(main_frame)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=15)

        # ✅ CONFIGURAR TREEVIEW COM NOVA COLUNA PRAZO
        columns = ("ID", "Nome", "Código", "Prazo")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=20)

        # Configurar colunas
        self.tree.heading("ID", text="ID")
        self.tree.heading("Nome", text="Nome do Fornecedor")
        self.tree.heading("Código", text="Código")
        self.tree.heading("Prazo", text="Prazo (dias)")  # ✅ NOVA COLUNA

        self.tree.column("ID", width=60, anchor="center")
        self.tree.column("Nome", width=600, anchor="w")
        self.tree.column("Código", width=100, anchor="center")
        self.tree.column("Prazo", width=120, anchor="center")  # ✅ NOVA COLUNA

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)

        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Pack treeview e scrollbars
        self.tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        v_scrollbar.pack(side="right", fill="y", pady=10)
        h_scrollbar.pack(side="bottom", fill="x", padx=(10, 0))

        # Bind eventos
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        if self.admin_mode:
            self.tree.bind("<Double-1>", lambda e: self.edit_supplier())

    def create_footer(self):
        """Cria rodapé com estatísticas"""
        footer_frame = ctk.CTkFrame(self.window, height=50)
        footer_frame.pack(fill="x", padx=20, pady=(0, 20))
        footer_frame.pack_propagate(False)

        self.stats_label = ctk.CTkLabel(
            footer_frame,
            text="Carregando estatísticas...",
            font=ctk.CTkFont(size=12)
        )
        self.stats_label.pack(expand=True)

        self.update_stats()

    def load_suppliers(self):
        """Carrega fornecedores na lista"""
        # Limpar lista atual
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Carregar fornecedores
        suppliers = self.db.get_all_suppliers()

        # Filtrar se há busca
        search_term = self.search_var.get().strip().lower()
        if search_term:
            suppliers = [
                s for s in suppliers
                if search_term in s.name.lower() or search_term in str(s.code)
            ]

        # ✅ ADICIONAR À LISTA COM PRAZO
        for supplier in suppliers:
            prazo_text = f"{supplier.prazo_dias}" if supplier.prazo_dias > 0 else "Não definido"
            self.tree.insert("", "end", values=(
                supplier.id,
                supplier.name,
                supplier.code,
                prazo_text
            ))

        self.update_stats()
        logger.info(f"Carregados {len(suppliers)} fornecedores")

    def on_search_change(self, *args):
        """Callback quando busca muda"""
        self.load_suppliers()

    def on_select(self, event):
        """Callback quando item é selecionado"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            values = item['values']

            # ✅ BUSCAR FORNECEDOR COMPLETO DO BANCO PARA TER PRAZO_DIAS
            supplier_id = int(values[0])
            self.selected_supplier = self.db.get_supplier_by_id(supplier_id)

            # ✅ HABILITAR BOTÕES APENAS NO MODO ADMIN
            if self.admin_mode:
                self.edit_btn.configure(state="normal")
                self.delete_btn.configure(state="normal")
        else:
            self.selected_supplier = None
            if self.admin_mode:
                self.edit_btn.configure(state="disabled")
                self.delete_btn.configure(state="disabled")

    def add_supplier(self):
        """Adiciona novo fornecedor"""
        if not self.admin_mode:
            messagebox.showwarning("Acesso Negado", "Função disponível apenas no modo administrador")
            return

        self.show_supplier_dialog()

    def edit_supplier(self):
        """Edita fornecedor selecionado"""
        if not self.admin_mode:
            messagebox.showwarning("Acesso Negado", "Função disponível apenas no modo administrador")
            return

        if not self.selected_supplier:
            messagebox.showwarning("Aviso", "Selecione um fornecedor para editar")
            return

        self.show_supplier_dialog(self.selected_supplier)

    def delete_supplier(self):
        """Exclui fornecedor selecionado"""
        if not self.admin_mode:
            messagebox.showwarning("Acesso Negado", "Função disponível apenas no modo administrador")
            return

        if not self.selected_supplier:
            messagebox.showwarning("Aviso", "Selecione um fornecedor para excluir")
            return

        if messagebox.askyesno(
                "Confirmar Exclusão",
                f"Tem certeza que deseja excluir o fornecedor:\n\n"
                f"Nome: {self.selected_supplier.name}\n"
                f"Código: {self.selected_supplier.code}\n"
                f"Prazo: {self.selected_supplier.prazo_dias} dias\n\n"
                f"Esta ação não pode ser desfeita!"
        ):
            if self.db.delete_supplier(self.selected_supplier.id):
                messagebox.showinfo("Sucesso", "Fornecedor excluído com sucesso!")
                self.load_suppliers()
            else:
                messagebox.showerror("Erro", "Erro ao excluir fornecedor")

    def import_suppliers(self):
        """Importa fornecedores de planilha"""
        try:
            file_path = filedialog.askopenfilename(
                title="Selecionar planilha de fornecedores",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )

            if not file_path:
                return

            # ✅ CORREÇÃO: Método agora retorna tupla
            sucessos, erros = self.db.import_from_spreadsheet(Path(file_path))

            if sucessos > 0:
                messagebox.showinfo(
                    "Importação Concluída",
                    f"Processados {sucessos} fornecedores com sucesso!\n"
                    f"(Incluindo novos e atualizações)\n"
                    f"Erros: {erros}"
                )
                self.refresh_list()
            else:
                messagebox.showerror(
                    "Erro na Importação",
                    f"Nenhum fornecedor foi processado.\n"
                    f"Erros: {erros}\n\n"
                    f"Verifique se a planilha tem as colunas:\n"
                    f"- Nome/Fornecedor\n"
                    f"- Código/ID\n"
                    f"- Prazo (opcional)"
                )

        except Exception as e:
            logger.error(f"Erro na importação: {e}")
            messagebox.showerror("Erro", f"Erro ao importar fornecedores:\n{e}")

    def show_import_preview(self, file_path: Path) -> bool:
        """Mostra preview dos dados antes da importação"""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            # Ler primeiras 10 linhas para preview
            preview_data = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), start=2):
                if len(row) >= 2 and row[0] and row[1]:
                    try:
                        name = str(row[0]).strip()
                        code = int(row[1])
                        # ✅ INCLUIR PRAZO NO PREVIEW SE DISPONÍVEL
                        prazo = int(row[2]) if len(row) > 2 and row[2] else 0
                        preview_data.append(f"Linha {row_num}: {name} → Código {code} → Prazo {prazo} dias")
                    except (ValueError, TypeError):
                        preview_data.append(f"Linha {row_num}: ERRO - {row[0]} → {row[1]}")
                elif row[0] or row[1]:  # Linha com dados parciais
                    preview_data.append(f"Linha {row_num}: ERRO - Dados incompletos")

            wb.close()

            if not preview_data:
                messagebox.showerror("Erro", "Nenhum dado válido encontrado no arquivo")
                return False

            # Mostrar preview
            preview_text = "\n".join(preview_data[:10])
            if len(preview_data) > 10:
                preview_text += f"\n... e mais {len(preview_data) - 10} linhas"

            result = messagebox.askyesno(
                "Preview da Importação",
                f"Dados encontrados no arquivo:\n\n{preview_text}\n\n"
                f"Deseja continuar com a importação?"
            )

            return result

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler arquivo:\n{e}")
            return False

    def create_example_file(self):
        """Cria arquivo de exemplo para importação"""
        if not self.admin_mode:
            messagebox.showwarning("Acesso Negado", "Função disponível apenas no modo administrador")
            return

        try:
            file_path = filedialog.asksaveasfilename(
                title="Salvar Arquivo de Exemplo",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialname="fornecedores_exemplo.xlsx"
            )

            if file_path:
                import openpyxl

                # Criar workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Fornecedores"

                # ✅ CABEÇALHOS COM PRAZO
                ws['A1'] = "Nome do Fornecedor"
                ws['B1'] = "Código"
                ws['C1'] = "Prazo (dias)"

                # ✅ DADOS DE EXEMPLO COM PRAZO
                example_data = [
                    ("DMOV", 51, 15),
                    ("D'Rossi Interiores", 1, 7),
                    ("Móveis Carraro", 2, 10),
                    ("Carraro Móveis", 2, 10),
                    ("Fábrica de Móveis Paulista", 3, 20),
                    ("Paulista Móveis", 3, 20),
                    ("Indústria de Móveis XYZ", 4, 14),
                    ("XYZ Móveis", 4, 14),
                    ("Casa & Decoração", 5, 12),
                    ("Decoração Casa Ltda", 5, 12),
                    ("Móveis Planejados ABC", 6, 25),
                    ("ABC Planejados", 6, 25),
                    ("Fornecedor Genérico", 0, 0),
                    ("Sem Fornecedor", 0, 0),
                ]

                # Inserir dados
                for row, (name, code, prazo) in enumerate(example_data, start=2):
                    ws[f'A{row}'] = name
                    ws[f'B{row}'] = code
                    ws[f'C{row}'] = prazo

                # Formatação
                from openpyxl.styles import Font, PatternFill

                # Cabeçalho em negrito
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                for col in ['A1', 'B1', 'C1']:
                    ws[col].font = header_font
                    ws[col].fill = header_fill

                # Ajustar largura das colunas
                ws.column_dimensions['A'].width = 35
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 15

                # Salvar
                wb.save(file_path)
                wb.close()

                messagebox.showinfo(
                    "Sucesso",
                    f"Arquivo de exemplo criado:\n{file_path}\n\n"
                    f"Use este arquivo como modelo para importar seus fornecedores.\n"
                    f"Inclui coluna de prazo de fabricação em dias."
                )

        except Exception as e:
            logger.error(f"Erro ao criar arquivo de exemplo: {e}")
            messagebox.showerror("Erro", f"Erro ao criar arquivo:\n{e}")

    def show_supplier_dialog(self, supplier: Optional[Supplier] = None):
        """Mostra diálogo de adição/edição"""
        dialog = SupplierDialog(self.window, supplier)
        self.window.wait_window(dialog.dialog)

        if dialog.result:
            # ✅ INCLUIR PRAZO NO RESULTADO
            name, code, prazo_dias = dialog.result

            if supplier:  # Editando
                if self.db.update_supplier(supplier.id, name, code, prazo_dias):
                    messagebox.showinfo("Sucesso", "Fornecedor atualizado com sucesso!")
                    self.load_suppliers()
                else:
                    messagebox.showerror("Erro",
                                         "Erro ao atualizar fornecedor.\nVerifique se nome e código não estão em uso.")
            else:  # Adicionando
                if self.db.add_supplier(name, code, prazo_dias):
                    messagebox.showinfo("Sucesso", "Fornecedor adicionado com sucesso!")
                    self.load_suppliers()
                else:
                    messagebox.showerror("Erro",
                                         "Erro ao adicionar fornecedor.\nVerifique se nome e código não estão em uso.")

    def export_suppliers(self):
        """Exporta fornecedores para JSON"""
        file_path = filedialog.asksaveasfilename(
            title="Salvar Dados de Fornecedores",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if file_path:
            if self.db.export_to_json(Path(file_path)):
                messagebox.showinfo("Sucesso", f"Dados exportados para:\n{file_path}")
            else:
                messagebox.showerror("Erro", "Erro ao exportar dados")

    def update_stats(self):
        """Atualiza estatísticas"""
        stats = self.db.get_statistics()  # ✅ CORRIGIDO
        current_count = len(self.tree.get_children())

        # ✅ INCLUIR ESTATÍSTICAS DE PRAZO
        if self.search_var.get().strip():
            text = f"📊 Mostrando {current_count} de {stats['total_suppliers']} fornecedores | 🔢 Códigos únicos: {stats['unique_codes']} | ⏱️ Com prazo: {stats['suppliers_with_prazo']}"
        else:
            text = f"📊 Total: {stats['total_suppliers']} fornecedores | 🔢 Códigos únicos: {stats['unique_codes']} | ⏱️ Com prazo: {stats['suppliers_with_prazo']} | 📈 Prazo médio: {stats['average_prazo_dias']} dias"

        self.stats_label.configure(text=text)


class SupplierDialog:
    """Diálogo para adicionar/editar fornecedor"""

    def __init__(self, parent, supplier: Optional[Supplier] = None):
        self.parent = parent
        self.supplier = supplier
        self.result = None

        self.setup_dialog()

    def setup_dialog(self):
        """Configura o diálogo"""
        self.dialog = ctk.CTkToplevel(self.parent)
        self.dialog.title("✏️ Fornecedor" if self.supplier else "➕ Novo Fornecedor")
        self.dialog.geometry("500x400")  # ✅ AUMENTAR ALTURA PARA NOVO CAMPO
        self.dialog.minsize(400, 350)

        # Centralizar
        self.dialog.transient(self.parent)
        self.dialog.grab_set()

        # Título
        title_text = "Editar Fornecedor" if self.supplier else "Novo Fornecedor"
        title_label = ctk.CTkLabel(
            self.dialog,
            text=title_text,
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title_label.pack(pady=(20, 30))

        # Formulário
        form_frame = ctk.CTkFrame(self.dialog)
        form_frame.pack(fill="x", padx=40, pady=(0, 20))

        # Nome
        ctk.CTkLabel(form_frame, text="Nome do Fornecedor:", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w",
                                                                                                            padx=20,
                                                                                                            pady=(20,
                                                                                                                  5))

        self.name_var = tk.StringVar(value=self.supplier.name if self.supplier else "")
        self.name_entry = ctk.CTkEntry(
            form_frame,
            textvariable=self.name_var,
            placeholder_text="Ex: D'Rossi Interiores",
            width=400
        )
        self.name_entry.pack(padx=20, pady=(0, 15))

        # Código
        ctk.CTkLabel(form_frame, text="Código:", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=20,
                                                                                                pady=(0, 5))

        self.code_var = tk.StringVar(value=str(self.supplier.code) if self.supplier else "")
        self.code_entry = ctk.CTkEntry(
            form_frame,
            textvariable=self.code_var,
            placeholder_text="Ex: 123",
            width=400
        )
        self.code_entry.pack(padx=20, pady=(0, 15))

        # ✅ NOVO CAMPO: PRAZO DE FABRICAÇÃO
        ctk.CTkLabel(form_frame, text="Prazo de Fabricação (dias):", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=20, pady=(0, 5))

        self.prazo_var = tk.StringVar(value=str(self.supplier.prazo_dias) if self.supplier else "0")
        self.prazo_entry = ctk.CTkEntry(
            form_frame,
            textvariable=self.prazo_var,
            placeholder_text="Ex: 15 (dias)",
            width=400
        )
        self.prazo_entry.pack(padx=20, pady=(0, 20))

        # ✅ INFORMAÇÃO SOBRE PRAZO
        info_label = ctk.CTkLabel(
            form_frame,
            text="ℹ️ Prazo em dias úteis para fabricação/entrega do produto",
            font=ctk.CTkFont(size=12),
            text_color=("gray60", "gray40")
        )
        info_label.pack(padx=20, pady=(0, 20))

        # Botões
        button_frame = ctk.CTkFrame(self.dialog, fg_color="transparent")
        button_frame.pack(fill="x", padx=40, pady=(0, 20))

        ctk.CTkButton(
            button_frame,
            text="💾 Salvar",
            command=self.save,
            width=120,
            height=40
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            button_frame,
            text="❌ Cancelar",
            command=self.cancel,
            width=120,
            height=40
        ).pack(side="left")

        # Focus no primeiro campo
        self.name_entry.focus()

    def save(self):
        """Salva fornecedor"""
        name = self.name_var.get().strip()
        code_str = self.code_var.get().strip()
        prazo_str = self.prazo_var.get().strip()

        # Validações
        if not name:
            messagebox.showerror("Erro", "Nome é obrigatório")
            return

        try:
            code = int(code_str)
            if code < 0:  # ✅ PERMITIR CÓDIGO 0
                raise ValueError()
        except ValueError:
            messagebox.showerror("Erro", "Código deve ser um número inteiro não negativo")
            return

        # ✅ VALIDAR PRAZO
        try:
            prazo_dias = int(prazo_str) if prazo_str else 0
            if prazo_dias < 0:
                raise ValueError()
        except ValueError:
            messagebox.showerror("Erro", "Prazo deve ser um número inteiro não negativo")
            return

        self.result = (name, code, prazo_dias)  # ✅ INCLUIR PRAZO NO RESULTADO
        self.dialog.destroy()

    def cancel(self):
        """Cancela operação"""
        self.dialog.destroy()