"""Escritor Excel otimizado"""

from pathlib import Path
from typing import List
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font # Importar estilos

from ..core.models import ProductDestination, VariationData, LojaWebData, KitData
from ..utils.logger import get_logger
from ..core.exceptions import ExcelProcessingError

logger = get_logger("excel_writer")

class ExcelWriter:
    """Escritor Excel otimizado"""

    def __init__(self):
        # ✅ MAPEAMENTO CORRIGIDO BASEADO NO TEMPLATE REAL
        self.column_mappings = {
            "PRODUTO": {
                "Código de Barras": "ean",
                "Código Fabricante": "cod_fabricante",
                "Fornecedor": "fornecedor",
                "Descrição Produto Nfe": "desc_nfe",
                "Descrição para Compra": "desc_compra",
                "Descrição Etiqueta": "desc_etiqueta",
                "Observação Produto": "obs_produto",
                "Complemento do Produto": "complemento_produto",
                "Categoria": "categoria",
                "Grupo": "grupo",
                "Cor do Produto": "cor",
                "Descrição para o Site": "desc_site",
                "Marca": "marca",
                "NCM": "ncm",

                # ✅ NOVOS CAMPOS DE PRECIFICAÇÃO
                "VR Custo Total": "vr_custo_total",
                "Custo IPI": "custo_ipi",
                "Custo Frete": "custo_frete",
                "Preço de Venda": "preco_de_venda",
                "Preço Promoção": "preco_promocao",

                "Fabricação Própria": "fabricacao_propria",
                "Tipo Produto": "tipo_produto",
                "Imprime Compl Pedido": "imprime_ped", # Faltou no ProductDestination, mas existe no mapping
                "Imprime Compl Compra": "imprime_comp", # Faltou no ProductDestination
                "Imprime Compl NF": "imprime_nf", # Faltou no ProductDestination
                "Site Marca": "site_marca",
                "Unidade de Venda": "unidade_venda", # Faltou no ProductDestination
                "Unidade de Compra": "unidade_compra", # Faltou no ProductDestination
                "Produto Inativo": "produto_inativo", # Faltou no ProductDestination
                "Dias de Garantia": "dias_garantia",
                "Site Garantia": "site_garantia",
                "Inicio Promoção": "inicio_promocao", # Faltou no ProductDestination
                "Fim Promoção": "fim_promocao", # Faltou no ProductDestination
                "Qtde Embalagem Venda": "qtde_emb_venda",
                "Qtde Volume": "qtde_volume",
                "Peso Bruto": "peso_bruto",
                "Peso Liquido": "peso_liquido",
                "Largura": "largura",
                "Altura": "altura",
                "Comprimento": "comprimento",
                "Diâmetro": "diametro",
                "Estoque Mínimo": "estoque_min",
                "Estoque de Segurança": "estoque_seg",
                "Dias para Entrega": "dias_entrega",
                "Site Disponibilidade": "site_disponibilidade",
                "Descrição HTML WEB": "desc_html"
            },
            "VARIACAO": {
                "EAN_FILHO": "ean_filho",
                "EAN_PAI": "ean_pai",
                "COR": "cor"
            },
            "LOJA WEB": {
                "EAN": "ean",
                "COD LOJA": "cod_loja",
                "Enviar para o Site": "enviar_site",
                "Disponibilizar Site": "disponibilizar_site",
                "Site Lançamento": "site_lancamento",
                "Site Destaque": "site_destaque",
                "CATEGORIA PRINCIPAL TRAY": "categoria_principal",
                "NIVEL ADICIONAL 1 TRAY": "nivel_1",
                "NIVEL ADICIONAL 2 TRAY": "nivel_2"
            },
            "KIT": {
                "EAN_KIT": "ean_kit",
                "EAN_COMPONENTE": "ean_componente",
                "QTDE": "quantidade",
                "% CUSTO DO KIT": "custo_kit",
                "% DESC VENDA": "desc_venda"
            }
        }

    def write_excel(
            self,
            output_path: Path,
            template_path: Path,
            produtos: List[ProductDestination],
            variacoes: List[VariationData],
            loja_web: List[LojaWebData],
            kits: List[KitData],
            origin_file: Path
    ):
        """Escreve dados no arquivo Excel usando template existente"""

        try:
            logger.info(f"Carregando template: {template_path}")

            if not template_path.exists():
                raise ExcelProcessingError(f"Template não encontrado: {template_path}")

            wb = openpyxl.load_workbook(template_path)
            logger.info(f"✅ Template carregado com abas: {wb.sheetnames}")

            # ✅ DEFINIR ESTILOS COM COR DE FUNDO E ALINHAMENTO
            # Cor de fundo verde claro
            fill_color = PatternFill(
                start_color="A9D08E",
                end_color="A9D08E",
                fill_type="solid"
            )

            # Alinhamento à esquerda
            alignment_left = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=False
            )

            # Font padrão
            font_default = Font(
                name="Calibri",
                size=11,
                bold=False
            )

            # ✅ FUNÇÃO PARA APLICAR ESTILO E FORMATO GERAL EM CÉLULAS PREENCHIDAS
            def apply_style_and_format_to_filled_cells(worksheet, start_row=2, sheet_name_for_mapping=""):
                """Aplica estilo, formato 'General' e trata Custo IPI=0 para células com conteúdo"""
                col_map = self._map_columns_from_template(worksheet, self.column_mappings[sheet_name_for_mapping])

                # Invert mapping for easier lookup by col_idx
                idx_to_header_map = {idx: header for header, idx in col_map.items()}

                for row in worksheet.iter_rows(min_row=start_row):
                    for cell in row:
                        # Verifica se a célula tem conteúdo
                        if cell.value is not None and str(cell.value).strip() != "":
                            cell.fill = fill_color
                            cell.alignment = alignment_left
                            cell.font = font_default

                            # ✅ SEMPRE DEFINIR FORMATO COMO 'GENERAL' PARA DADOS
                            cell.number_format = 'General'

                            # ✅ TRATAMENTO ESPECIAL PARA 'Custo IPI' = 0
                            header = idx_to_header_map.get(cell.column)
                            if header == "Custo IPI" and (cell.value == 0 or cell.value == 0.0):
                                cell.value = 0 # Garante que o valor numérico é 0
                                cell.number_format = 'General' # Explicita o formato geral

            # 1. Processa abas com lógica
            self._write_produtos(wb, produtos)
            # ✅ APLICAR ESTILO E FORMATO NA ABA PRODUTO
            if "PRODUTO" in wb.sheetnames and produtos:
                apply_style_and_format_to_filled_cells(wb["PRODUTO"], start_row=2, sheet_name_for_mapping="PRODUTO")
                logger.info(f"  🎨 Estilo e formato aplicados em {len(produtos)} produtos na aba PRODUTO")

            self._write_variacoes(wb, variacoes)
            # ✅ APLICAR ESTILO E FORMATO NA ABA VARIACAO
            if "VARIACAO" in wb.sheetnames and variacoes:
                apply_style_and_format_to_filled_cells(wb["VARIACAO"], start_row=2, sheet_name_for_mapping="VARIACAO")
                logger.info(f"  🎨 Estilo e formato aplicados em {len(variacoes)} variações na aba VARIACAO")

            self._write_loja_web(wb, loja_web)
            # ✅ APLICAR ESTILO E FORMATO NA ABA LOJA WEB
            if "LOJA WEB" in wb.sheetnames and loja_web:
                apply_style_and_format_to_filled_cells(wb["LOJA WEB"], start_row=2, sheet_name_for_mapping="LOJA WEB")
                logger.info(f"  🎨 Estilo e formato aplicados em {len(loja_web)} itens na aba LOJA WEB")

            self._write_kits(wb, kits)
            # ✅ APLICAR ESTILO E FORMATO NA ABA KIT
            if "KIT" in wb.sheetnames and kits:
                apply_style_and_format_to_filled_cells(wb["KIT"], start_row=2, sheet_name_for_mapping="KIT")
                logger.info(f"  🎨 Estilo e formato aplicados em {len(kits)} kits na aba KIT")

            # 2. Copia abas da origem sem alteração
            self._copy_sheets_from_origin(wb, origin_file)

            # Salva arquivo
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            logger.success(f"✅ Arquivo salvo: {output_path}")
            logger.success(f"🎨 Estilo aplicado: Cor de fundo #A9D08E, alinhamento à esquerda e FORMATO 'GERAL' para dados.")

        except Exception as e:
            logger.error(f"Erro ao escrever Excel: {e}")
            raise ExcelProcessingError(f"Falha ao criar arquivo: {e}")

    def _write_produtos(self, wb: openpyxl.Workbook, produtos: List[ProductDestination]):
        """Escreve aba PRODUTO"""
        sheet_name = "PRODUTO"

        print(f"\n📝 === DEBUG WRITER _write_produtos ===")
        print(f"📊 Produtos recebidos: {len(produtos)}")
        print(f"📊 Tipo da lista: {type(produtos)}")

        if produtos:
            print(f"✅ Primeiro produto:")
            primeiro = produtos[0]
            print(f"  - EAN: {primeiro.ean}")
            print(f"  - Tipo: {type(primeiro)}")
            print(f"  - Categoria (Cat. da origem): {primeiro.categoria}")
            print(f"  - Grupo: {primeiro.grupo}")
            print(f"  - Cor: {primeiro.cor}")
            print(f"  - Desc Site: {primeiro.desc_site}")
        else:
            print("❌ LISTA DE PRODUTOS ESTÁ VAZIA NO WRITER!")

        print(f"=" * 50)

        if sheet_name not in wb.sheetnames:
            raise ExcelProcessingError(f"Aba '{sheet_name}' não encontrada no template")

        ws = wb[sheet_name]
        logger.info(f"📝 Processando aba {sheet_name}...")

        # Mapeia colunas existentes
        col_map = self._map_columns_from_template(ws, self.column_mappings[sheet_name])
        logger.info(f"�� Colunas mapeadas: {len(col_map)}/{len(self.column_mappings[sheet_name])}")

        # ✅ DEBUG: Verificar dados do primeiro produto (mantido para contexto)
        if produtos:
            primeiro_produto = produtos[0]
            logger.info("�� === DADOS DO PRIMEIRO PRODUTO ===")
            logger.info(f"  EAN: '{primeiro_produto.ean}'")
            logger.info(f"  Categoria (Cat. da origem): '{primeiro_produto.categoria}'")
            logger.info(f"  Grupo: '{primeiro_produto.grupo}'")
            logger.info(f"  Cor: '{primeiro_produto.cor}'")
            logger.info(f"  Desc Site: '{primeiro_produto.desc_site}'")
            logger.info(f"  Site Garantia: '{primeiro_produto.site_garantia}'")
            logger.info(f"  Desc HTML: '{primeiro_produto.desc_html[:100] if primeiro_produto.desc_html else None}...'")

            logger.info("💰 === DADOS DE PRECIFICAÇÃO ===")
            logger.info(f"  VR Custo Total: R\$ {primeiro_produto.vr_custo_total:.2f}")
            logger.info(f"  Custo IPI: R\$ {primeiro_produto.custo_ipi:.2f}")
            logger.info(f"  Custo Frete: R\$ {primeiro_produto.custo_frete:.2f}")
            logger.info(f"  Preço de Venda: R\$ {primeiro_produto.preco_de_venda:.2f}")
            logger.info(f"  Preço Promoção: R\$ {primeiro_produto.preco_promocao:.2f}")

        # Limpa dados existentes (mantém cabeçalho na linha 1)
        self._clear_data_rows(ws, start_row=2)

        # Escreve novos dados começando da linha 2
        for row_idx, produto in enumerate(produtos, start=2):
            for header, attr in self.column_mappings[sheet_name].items():
                if header in col_map:
                    col_idx = col_map[header]
                    value = getattr(produto, attr, "")

                    # ✅ Escreve o valor na célula
                    ws.cell(row=row_idx, column=col_idx, value=value)

                    # O apply_style_and_format_to_filled_cells é chamado depois para formatar,
                    # então não precisamos setar o formato aqui diretamente para evitar duplicação.
                    # Mas podemos ter logs específicos para garantir que os valores estão ok.

                    # ✅ DEBUG ESPECÍFICO PARA CAMPOS PROBLEMÁTICOS E PRECIFICAÇÃO
                    if attr in ["site_garantia", "desc_site", "cor", "desc_html", "categoria", "grupo", "vr_custo_total", "custo_ipi",
                                "custo_frete", "preco_de_venda", "preco_promocao", "qtde_volume", "peso_bruto", "largura", "altura", "comprimento", "estoque_seg", "dias_entrega", "site_disponibilidade"] and row_idx == 2:
                        if attr in ["vr_custo_total", "custo_ipi", "custo_frete", "preco_de_venda", "preco_promocao"]:
                            logger.info(f"  💰 Escrevendo {attr}: R\$ {value:.2f} na coluna {col_idx} ('{header}')")
                        else:
                            logger.info(f"  📝 Escrevendo {attr}: '{value}' na coluna {col_idx} ('{header}')")

        logger.success(f"✅ {len(produtos)} produtos escritos na aba PRODUTO")

    def _write_variacoes(self, wb: openpyxl.Workbook, variacoes: List[VariationData]):
        """Escreve aba VARIACAO"""
        sheet_name = "VARIACAO"

        if sheet_name not in wb.sheetnames:
            logger.warning(f"Aba '{sheet_name}' não encontrada no template")
            return

        ws = wb[sheet_name]
        col_map = self._map_columns_from_template(ws, self.column_mappings[sheet_name])
        self._clear_data_rows(ws, start_row=2)

        for row_idx, variacao in enumerate(variacoes, start=2):
            for header, attr in self.column_mappings[sheet_name].items():
                if header in col_map:
                    col_idx = col_map[header]
                    value = getattr(variacao, attr, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

        logger.info(f"✅ {len(variacoes)} variações escritas na aba VARIACAO")

    def _write_loja_web(self, wb: openpyxl.Workbook, loja_web: List[LojaWebData]):
        """Escreve aba LOJA WEB"""
        sheet_name = "LOJA WEB"

        if sheet_name not in wb.sheetnames:
            logger.warning(f"Aba '{sheet_name}' não encontrada no template")
            return

        ws = wb[sheet_name]
        col_map = self._map_columns_from_template(ws, self.column_mappings[sheet_name])
        self._clear_data_rows(ws, start_row=2)

        for row_idx, loja in enumerate(loja_web, start=2):
            for header, attr in self.column_mappings[sheet_name].items():
                if header in col_map:
                    col_idx = col_map[header]
                    value = getattr(loja, attr, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

        logger.info(f"✅ {len(loja_web)} registros escritos na aba LOJA WEB")

    def _write_kits(self, wb: openpyxl.Workbook, kits: List[KitData]):
        """Escreve aba KIT"""
        sheet_name = "KIT"

        if sheet_name not in wb.sheetnames:
            logger.warning(f"Aba '{sheet_name}' não encontrada no template")
            return

        ws = wb[sheet_name]
        col_map = self._map_columns_from_template(ws, self.column_mappings[sheet_name])
        self._clear_data_rows(ws, start_row=2)

        for row_idx, kit in enumerate(kits, start=2):
            for header, attr in self.column_mappings[sheet_name].items():
                if header in col_map:
                    col_idx = col_map[header]
                    value = getattr(kit, attr, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

        logger.info(f"✅ {len(kits)} kits escritos na aba KIT")

    def _copy_sheets_from_origin(self, wb_dest: openpyxl.Workbook, origin_file: Path):
        """Copia abas específicas da origem sem alteração"""

        sheets_to_copy = [
            "Instruções de Preenchimento",
            "Tipo Importação"
        ]

        try:
            if not origin_file.exists():
                logger.warning(f"Arquivo de origem não encontrado para copiar abas: {origin_file}")
                return

            # Carrega workbook de origem
            wb_origin = openpyxl.load_workbook(origin_file, read_only=False)
            logger.info(f"Abas disponíveis na origem: {wb_origin.sheetnames}")

            for sheet_name in sheets_to_copy:
                if sheet_name in wb_origin.sheetnames:
                    try:
                        # Remove aba existente no destino se houver
                        if sheet_name in wb_dest.sheetnames:
                            wb_dest.remove(wb_dest[sheet_name])
                            logger.info(f"Aba '{sheet_name}' removida do destino")

                        # Copia aba da origem
                        ws_origin = wb_origin[sheet_name]
                        ws_dest = wb_dest.create_sheet(sheet_name)

                        # Copia todos os dados e formatação
                        for row in ws_origin.iter_rows():
                            for cell in row:
                                new_cell = ws_dest[cell.coordinate]
                                new_cell.value = cell.value

                                # Copia formatação se necessário
                                if cell.has_style:
                                    try:
                                        new_cell.font = cell.font.copy()
                                        new_cell.border = cell.border.copy()
                                        new_cell.fill = cell.fill.copy()
                                        new_cell.number_format = cell.number_format
                                        new_cell.protection = cell.protection.copy()
                                        new_cell.alignment = cell.alignment.copy()
                                    except Exception as style_error:
                                        logger.debug(f"Erro ao copiar estilo da célula {cell.coordinate}: {style_error}")
                        # Copia dimensões das colunas
                        for col_letter, col_dim in ws_origin.column_dimensions.items():
                            ws_dest.column_dimensions[col_letter].width = col_dim.width

                        # Copia dimensões das linhas
                        for row_num, row_dim in ws_origin.row_dimensions.items():
                            ws_dest.row_dimensions[row_num].height = row_dim.height

                        logger.info(f"✅ Aba '{sheet_name}' copiada da origem")

                    except Exception as copy_error:
                        logger.error(f"Erro ao copiar aba '{sheet_name}': {copy_error}")
                else:
                    logger.warning(f"Aba '{sheet_name}' não encontrada na origem")

            wb_origin.close()

        except Exception as e:
            logger.error(f"Erro ao copiar abas da origem: {e}")
            # Não falha o processo, apenas avisa

    def _map_columns_from_template(self, ws: Worksheet, expected_mapping: dict) -> dict:
        """Mapeia colunas do template existente"""
        col_map = {}

        # Procura na linha 1 (cabeçalhos)
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if header_value:
                header_clean = str(header_value).strip()
                if header_clean in expected_mapping:
                    col_map[header_clean] = col
                    logger.debug(f"✅ Mapeado: '{header_clean}' → Coluna {col}")

        # Verifica se encontrou colunas essenciais
        if not col_map:
            available_headers = []
            for col in range(1, min(ws.max_column + 1, 20)):  # Primeiras 20 colunas
                header = ws.cell(row=1, column=col).value
                if header:
                    available_headers.append(str(header).strip())

            raise ExcelProcessingError(
                f"Nenhuma coluna reconhecida no template. "
                f"Cabeçalhos encontrados: {available_headers[:10]}..."
            )

        return col_map

    def _clear_data_rows(self, ws: Worksheet, start_row: int = 2):
        """Limpa dados das linhas (mantém cabeçalhos na linha 1)"""
        if ws.max_row < start_row:
            return

        # Limpa todas as células de dados (linha 2 em diante)
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col, value=None)

        logger.debug(f"Dados limpos da linha {start_row} até {ws.max_row}")
