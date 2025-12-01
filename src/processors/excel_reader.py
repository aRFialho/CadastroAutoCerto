import polars as pl
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from loguru import logger
from ..core.models import ProductOrigin
from ..core.exceptions import ExcelProcessingError


class ExcelReader:
    """Leitor Excel otimizado com Polars + OpenPyXL"""

    def __init__(self):
        self.column_mapping = {
            # ✅ CORREÇÃO: EAN agora é "EAN e Variação"
            "EAN": "ean",  # Fallback
            "EAN e Variação": "ean",  # ← NOVO NOME
            "Ean e Variação": "ean",  # ← VARIAÇÃO
            "ean e variação": "ean",  # ← VARIAÇÃO
            "EAN VARIAÇÃO": "ean_variacao",
            "COD. FORNECEDOR": "cod_fornecedor",
            "CÓD. FORNECEDOR": "cod_fornecedor",
            "Cod. Fornecedor": "cod_fornecedor",
            "Cód. Fornecedor": "cod_fornecedor",
            "Complemento/Título Interno": "complemento_titulo",

            # ✅ ALTERAÇÃO: Categoria original (para LOJA WEB)
            "CATEGORIA": "categoria",
            "Categoria": "categoria",
            "categoria": "categoria",

            # ✅ NOVO: Mapeamento para "Cat." (para aba PRODUTO)
            "CAT.": "cat",
            "Cat.": "cat",
            "cat.": "cat",
            "CAT": "cat",
            "Cat": "cat",
            "cat": "cat",

            # ✅ MANTER: Grupo
            "GRUPO": "grupo",
            "Grupo": "grupo",
            "grupo": "grupo",

            # ✅ CORREÇÃO: Nova coluna "Título para Compra"
            "Título para Compra": "titulo_compra",  # ← NOVA COLUNA

            # ✅ CORREÇÃO: Todas as variações de ANÚNCIO
            "ANÚNCIO": "anuncio",
            "Anúncio": "anuncio",
            "anúncio": "anuncio",
            "ANUNCIO": "anuncio",
            "Anuncio": "anuncio",
            "anuncio": "anuncio",

            # ✅ OPCIONAL: Produto não será mais usado para preenchimento
            "PRODUTO": "produto",
            "Produto": "produto",

            # ✅ CORREÇÃO PRINCIPAL: TIPO DE PRODUTO
            "Tipo de Produto": "tipo_produto",  # ← NOME PRINCIPAL
            "Tipo do Produto": "tipo_produto",  # ← VARIAÇÃO ENCONTRADA
            "TIPO DE PRODUTO": "tipo_produto",  # ← MAIÚSCULA
            "TIPO DO PRODUTO": "tipo_produto",  # ← VARIAÇÃO
            "tipo de produto": "tipo_produto",  # ← MINÚSCULA
            "tipo do produto": "tipo_produto",  # ← VARIAÇÃO
            "Tipo De Produto": "tipo_produto",  # ← TITLE CASE

            # ✅ CORREÇÃO PRINCIPAL: "Cor do Produto" - TODAS AS VARIAÇÕES
            "Cor do Produto": "cor",  # ← NOME PRINCIPAL
            "COR DO PRODUTO": "cor",  # ← MAIÚSCULA
            "cor do produto": "cor",  # ← MINÚSCULA
            "Cor Do Produto": "cor",  # ← TITLE CASE
            "COR": "cor",  # ← FALLBACK
            "Cor": "cor",  # ← FALLBACK
            "cor": "cor",  # ← FALLBACK
            "Cor do tecido": "cor",  # ← FALLBACK ANTIGO
            "COR DO TECIDO": "cor",  # ← FALLBACK ANTIGO
            "cor do tecido": "cor",  # ← FALLBACK ANTIGO
            "Cor do Tecido": "cor",  # ← CASE ENCONTRADO NOS LOGS

            "Tipo de anúncio": "tipo_anuncio",
            "VOLUMES": "volumes",

            # Dimensões
            "Qtde Volume": "qtde_volume",
            "Peso Bruto": "peso_bruto",
            "Largura": "largura",
            "Altura": "altura",
            "Comprimento": "comprimento",

            # Fallbacks
            "PESO BRUTO": "peso_bruto",
            "PESO LÍQUIDO": "peso_liquido",
            "EMB.LARGURA": "largura",
            "EMB.ALTURA": "altura",
            "EMB.COMPRIMENTO": "comprimento",

            "PRAZO": "prazo",
            "DESCRIÇÃO HTML": "descricao_html",
            "NCM": "ncm",

            # ✅ CORREÇÃO: Complemento do Produto com novo nome
            "Complemento do Produto": "complemento_produto",  # Fallback
            "Complemento do Produto (sem o cod e marca)": "complemento_produto",  # ← NOVO NOME
            "COMPLEMENTO DO PRODUTO (SEM O COD E MARCA)": "complemento_produto",  # ← VARIAÇÃO
            "complemento do produto (sem o cod e marca)": "complemento_produto"  # ← VARIAÇÃO
        }

    def read_products(self, file_path: Path, sheet_name: str = "Produtos") -> List[ProductOrigin]:
        """Lê produtos da planilha de origem"""
        try:
            logger.info(f"Lendo planilha: {file_path} - Aba: {sheet_name}")

            # ✅ ESTRATÉGIA MAIS ROBUSTA: PANDAS PURO + CONVERSÃO MANUAL
            try:
                import pandas as pd

                # Ler com pandas
                df_pandas = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')

                # Verificar se DataFrame não está vazio
                if df_pandas.empty:
                    raise ExcelProcessingError("Planilha está vazia")

                # ✅ LIMPAR NOMES DAS COLUNAS RIGOROSAMENTE
                original_columns = df_pandas.columns.tolist()
                cleaned_columns = []

                logger.info(f"🔍 === LIMPANDO NOMES DAS COLUNAS ===")
                for i, col in enumerate(original_columns):
                    if col is None or pd.isna(col):
                        new_name = f"Coluna_Sem_Nome_{i + 1}"
                        logger.warning(f"⚠️ Coluna {i + 1} sem nome, renomeando para: '{new_name}'")
                        cleaned_columns.append(new_name)
                    else:
                        # Converter para string e limpar rigorosamente
                        col_str = str(col).strip()
                        if col_str == "" or col_str.lower() in ["unnamed", "nan", "none"]:
                            new_name = f"Coluna_Vazia_{i + 1}"
                            logger.warning(f"⚠️ Coluna {i + 1} vazia, renomeando para: '{new_name}'")
                            cleaned_columns.append(new_name)
                        else:
                            # Remover caracteres problemáticos
                            clean_name = col_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
                            clean_name = ' '.join(clean_name.split())  # Remove espaços extras
                            cleaned_columns.append(clean_name)
                            logger.info(f"✅ Coluna {i + 1}: '{col}' → '{clean_name}'")

                # Aplicar nomes limpos
                df_pandas.columns = cleaned_columns

                # ✅ CONVERSÃO MANUAL PARA POLARS (SEM PYARROW)
                # Converter tipos problemáticos para tipos simples
                for col in df_pandas.columns:
                    if df_pandas[col].dtype == 'object':
                        # Converter object para string
                        df_pandas[col] = df_pandas[col].astype(str)
                    elif df_pandas[col].dtype.name.startswith('Int'):
                        # Converter Int64 nullable para int64 simples
                        df_pandas[col] = df_pandas[col].fillna(0).astype('int64')
                    elif df_pandas[col].dtype.name.startswith('Float'):
                        # Converter Float64 nullable para float64 simples
                        df_pandas[col] = df_pandas[col].fillna(0.0).astype('float64')

                # ✅ AGORA CONVERTER PARA POLARS
                try:
                    df = pl.from_pandas(df_pandas)
                    logger.info("✅ Conversão Pandas → Polars bem-sucedida")
                except Exception as conv_error:
                    logger.error(f"❌ Erro na conversão: {conv_error}")
                    # Fallback: criar DataFrame Polars manualmente
                    data_dict = {}
                    for col in df_pandas.columns:
                        data_dict[col] = df_pandas[col].tolist()
                    df = pl.DataFrame(data_dict)
                    logger.info("✅ Conversão manual para Polars bem-sucedida")

                logger.info(f"Planilha carregada: {df.shape[0]} linhas, {df.shape[1]} colunas")
                logger.info(f"Colunas encontradas: {df.columns}")

            except Exception as read_error:
                logger.error(f"❌ Erro na leitura com Pandas: {read_error}")
                raise ExcelProcessingError(f"Falha na leitura do arquivo: {read_error}")

            # Mapeia colunas
            df = self._map_columns(df)

            # Limpa e valida dados
            df = self._clean_data(df)

            # DEBUG: Mostra algumas linhas para verificar dados importantes
            logger.info("🔍 Primeiras 5 linhas após limpeza:")
            try:
                for i, row in enumerate(df.head(5).iter_rows(named=True)):
                    categoria = row.get('categoria', 'N/A')  # Para LOJA WEB
                    cat = row.get('cat', 'N/A')  # Para aba PRODUTO
                    grupo = row.get('grupo', 'N/A')
                    tipo_prod = row.get('tipo_produto', 'N/A')
                    cor = row.get('cor', 'N/A')
                    anuncio = row.get('anuncio', 'N/A')
                    logger.info(
                        f"Linha {i + 1}: EAN={row.get('ean')}, CAT.='{cat}', CATEGORIA='{categoria}', GRUPO='{grupo}', TIPO='{tipo_prod}', COR='{cor}', ANÚNCIO='{anuncio}'"
                    )
            except Exception as debug_error:
                logger.warning(f"⚠️ Erro no debug das primeiras linhas: {debug_error}")

            # Converte para modelos Pydantic
            products = []
            for row in df.iter_rows(named=True):
                try:
                    # Converte None para valores padrão
                    clean_row = self._clean_row_data(row)
                    product = ProductOrigin(**clean_row)
                    products.append(product)
                except Exception as e:
                    logger.warning(f"Erro ao processar linha: {row}. Erro: {e}")

            # DEBUG: Estatísticas importantes
            with_categoria = [p for p in products if p.categoria and str(p.categoria).strip()]  # Para LOJA WEB
            with_cat = [p for p in products if p.cat and str(p.cat).strip()]  # Para aba PRODUTO
            with_grupo = [p for p in products if p.grupo and str(p.grupo).strip()]
            with_color = [p for p in products if p.cor and str(p.cor).strip()]
            with_tipo_produto = [p for p in products if p.tipo_produto and str(p.tipo_produto).strip()]

            logger.info(f"📊 Estatísticas:")
            logger.info(f"  - Total produtos: {len(products)}")
            logger.info(f"  - Com CATEGORIA definida (LOJA WEB): {len(with_categoria)}")
            logger.info(f"  - Com CAT. definida (aba PRODUTO): {len(with_cat)}")
            logger.info(f"  - Com GRUPO definido: {len(with_grupo)}")
            logger.info(f"  - Com COR DO PRODUTO definida: {len(with_color)}")
            logger.info(f"  - Com TIPO DE PRODUTO definido: {len(with_tipo_produto)}")

            # ✅ MOSTRA CATEGORIAS E CAT. ENCONTRADAS
            if with_cat:
                cat_unicas = list(set([p.cat for p in with_cat if p.cat]))
                logger.info(f"🏷️ Cat. encontradas (aba PRODUTO): {cat_unicas[:10]}...")  # Primeiras 10

            if with_categoria:
                categorias_unicas = list(set([p.categoria for p in with_categoria if p.categoria]))
                logger.info(f"🏷️ Categorias encontradas (LOJA WEB): {categorias_unicas[:10]}...")  # Primeiras 10

            logger.success(f"✅ {len(products)} produtos carregados com sucesso")
            # ✅ ADICIONAR ANTES DO RETURN NO FINAL DO MÉTODO read_products
            logger.info(f"📋 === DEBUG READER read_products ===")
            logger.info(f"📊 Produtos finais sendo retornados: {len(products)}")
            logger.info(f"📊 Tipo da lista: {type(products)}")
            if products:
                logger.info(f"📊 Primeiro produto: EAN={products[0].ean}, TIPO={products[0].tipo_produto}")
                logger.info(f"📊 Último produto: EAN={products[-1].ean}, TIPO={products[-1].tipo_produto}")
            else:
                logger.error("❌ LISTA DE PRODUTOS VAZIA NO READER!")
            logger.info("=" * 50)

            return products

        except Exception as e:
            logger.error(f"❌ Erro ao ler planilha: {e}")
            raise ExcelProcessingError(f"Falha ao ler {file_path}: {e}")

    def _normalize_column_name(self, col_name: str) -> str:
        """Normaliza nome da coluna removendo acentos e espaços"""
        if not col_name:
            return ""

        # Remove acentos comuns
        normalized = col_name.lower().strip()
        normalized = normalized.replace("ú", "u").replace("ã", "a").replace("ç", "c")
        normalized = normalized.replace("á", "a").replace("é", "e").replace("í", "i")
        normalized = normalized.replace("ó", "o").replace("ê", "e").replace("ô", "o")
        normalized = normalized.replace(" ", "").replace(".", "").replace("_", "")
        normalized = normalized.replace("(", "").replace(")", "")

        return normalized

    def _map_columns(self, df: pl.DataFrame) -> pl.DataFrame:
        """Mapeia nomes das colunas com busca inteligente"""
        rename_dict = {}

        # 🔍 DEBUG: Mostra todas as colunas encontradas
        logger.info(f"🔍 === COLUNAS DISPONÍVEIS NA PLANILHA DE ORIGEM ===")
        for i, col in enumerate(df.columns):
            logger.info(f"  {i + 1:2d}: '{col}'")

            # 🎯 BUSCA ESPECÍFICA POR NOVAS COLUNAS
            col_lower = col.lower()
            if "ean" in col_lower and "variação" in col_lower:
                logger.warning(f"🎯 COLUNA EAN E VARIAÇÃO ENCONTRADA: '{col}'")
            elif "tipo" in col_lower and "produto" in col_lower:
                logger.warning(f"�� COLUNA TIPO DE PRODUTO ENCONTRADA: '{col}'")
            elif "cor" in col_lower and "produto" in col_lower:
                logger.warning(f"🎯 COLUNA COR DO PRODUTO ENCONTRADA: '{col}'")
            elif "cor" in col_lower and "tecido" in col_lower:
                logger.warning(f"🎯 COLUNA COR DO TECIDO ENCONTRADA (FALLBACK): '{col}'")
            elif "cor" in col_lower and len(col_lower.strip()) <= 4:  # Só "COR" ou "Cor"
                logger.warning(f"🎯 COLUNA COR SIMPLES ENCONTRADA: '{col}'")
            elif "título" in col_lower and "compra" in col_lower:
                logger.warning(f"🎯 COLUNA TÍTULO PARA COMPRA ENCONTRADA: '{col}'")
            elif "complemento" in col_lower and "sem" in col_lower:
                logger.warning(f"🎯 COLUNA COMPLEMENTO (SEM COD E MARCA) ENCONTRADA: '{col}'")
            elif self._is_anuncio_column(col):
                logger.warning(f"🎯 COLUNA DE ANÚNCIO ENCONTRADA: '{col}'")
            elif "categoria" in col_lower:
                logger.warning(f"🎯 COLUNA CATEGORIA ENCONTRADA (LOJA WEB): '{col}'")
            elif "cat" in col_lower and len(col_lower.strip()) <= 5:  # "Cat." ou "CAT"
                logger.warning(f"🎯 COLUNA CAT. ENCONTRADA (aba PRODUTO): '{col}'")
            elif "grupo" in col_lower:
                logger.warning(f"🎯 COLUNA GRUPO ENCONTRADA: '{col}'")

        # ✅ MAPEAMENTO DIRETO PRIMEIRO (MAIS ESPECÍFICO PRIMEIRO)
        for col in df.columns:
            col_clean = col.strip() if col else ""
            if col_clean in self.column_mapping:
                rename_dict[col] = self.column_mapping[col_clean]
                logger.info(f"✅ Mapeamento direto: '{col_clean}' -> '{self.column_mapping[col_clean]}'")

        # ✅ BUSCA INTELIGENTE PARA COLUNAS NÃO MAPEADAS
        mapped_fields = set(rename_dict.values())

        for col in df.columns:
            if col not in rename_dict:  # Se ainda não foi mapeada
                col_normalized = self._normalize_column_name(col)

                # ✅ BUSCA POR TIPO DE PRODUTO (PRIORIDADE MÁXIMA)
                if "tipo_produto" not in mapped_fields and self._is_tipo_produto_column(col):
                    rename_dict[col] = "tipo_produto"
                    logger.warning(f"🔄 Mapeamento inteligente TIPO DE PRODUTO: '{col}' -> 'tipo_produto'")
                    mapped_fields.add("tipo_produto")

                # Busca por EAN E VARIAÇÃO
                elif "ean" not in mapped_fields and self._is_ean_variacao_column(col):
                    rename_dict[col] = "ean"
                    logger.warning(f"�� Mapeamento inteligente EAN: '{col}' -> 'ean'")
                    mapped_fields.add("ean")

                # ✅ BUSCA POR COR DO PRODUTO (PRIORIDADE MÁXIMA)
                elif "cor" not in mapped_fields and self._is_cor_produto_column(col):
                    rename_dict[col] = "cor"
                    logger.warning(f"🔄 Mapeamento inteligente COR DO PRODUTO: '{col}' -> 'cor'")
                    mapped_fields.add("cor")

                # ✅ BUSCA POR COR SIMPLES (SEGUNDA PRIORIDADE)
                elif "cor" not in mapped_fields and self._is_cor_simples_column(col):
                    rename_dict[col] = "cor"
                    logger.warning(f"🔄 Mapeamento inteligente COR SIMPLES: '{col}' -> 'cor'")
                    mapped_fields.add("cor")

                # ✅ BUSCA POR COR DO TECIDO (FALLBACK)
                elif "cor" not in mapped_fields and self._is_cor_tecido_column(col):
                    rename_dict[col] = "cor"
                    logger.warning(f"🔄 Mapeamento inteligente COR DO TECIDO (FALLBACK): '{col}' -> 'cor'")
                    mapped_fields.add("cor")

                # Busca por TÍTULO PARA COMPRA
                elif "titulo_compra" not in mapped_fields and self._is_titulo_compra_column(col):
                    rename_dict[col] = "titulo_compra"
                    logger.warning(f"🔄 Mapeamento inteligente TÍTULO COMPRA: '{col}' -> 'titulo_compra'")
                    mapped_fields.add("titulo_compra")

                # Busca por COMPLEMENTO (SEM COD E MARCA)
                elif "complemento_produto" not in mapped_fields and self._is_complemento_sem_cod_column(col):
                    rename_dict[col] = "complemento_produto"
                    logger.warning(f"🔄 Mapeamento inteligente COMPLEMENTO: '{col}' -> 'complemento_produto'")
                    mapped_fields.add("complemento_produto")

                # Busca por ANÚNCIO
                elif "anuncio" not in mapped_fields and self._is_anuncio_column(col):
                    rename_dict[col] = "anuncio"
                    logger.warning(f"�� Mapeamento inteligente ANÚNCIO: '{col}' -> 'anuncio'")
                    mapped_fields.add("anuncio")

                # ✅ BUSCA POR CATEGORIA (LOJA WEB)
                elif "categoria" not in mapped_fields and self._is_categoria_column(col):
                    rename_dict[col] = "categoria"
                    logger.warning(f"🔄 Mapeamento inteligente CATEGORIA (LOJA WEB): '{col}' -> 'categoria'")
                    mapped_fields.add("categoria")

                # ✅ NOVO: BUSCA POR CAT. (aba PRODUTO)
                elif "cat" not in mapped_fields and self._is_cat_column(col):
                    rename_dict[col] = "cat"
                    logger.warning(f"🔄 Mapeamento inteligente CAT. (aba PRODUTO): '{col}' -> 'cat'")
                    mapped_fields.add("cat")

                # ✅ BUSCA POR GRUPO
                elif "grupo" not in mapped_fields and self._is_grupo_column(col):
                    rename_dict[col] = "grupo"
                    logger.warning(f"🔄 Mapeamento inteligente GRUPO: '{col}' -> 'grupo'")
                    mapped_fields.add("grupo")

        # ✅ VERIFICAÇÕES FINAIS
        if "tipo_produto" not in mapped_fields:
            logger.error("🚨 ATENÇÃO: Nenhuma coluna de TIPO DE PRODUTO foi mapeada!")
            logger.error("🚨 Colunas disponíveis que contêm 'tipo':")
            for col in df.columns:
                if "tipo" in col.lower():
                    logger.error(f"    - '{col}'")

        if "cor" not in mapped_fields:
            logger.error("🚨 ATENÇÃO: Nenhuma coluna de COR foi mapeada!")
            logger.error("🚨 Colunas disponíveis que contêm 'cor':")
            for col in df.columns:
                if "cor" in col.lower():
                    logger.error(f"    - '{col}'")

        if "categoria" not in mapped_fields:
            logger.warning("⚠️ ATENÇÃO: Nenhuma coluna de CATEGORIA (LOJA WEB) foi mapeada!")

        if "cat" not in mapped_fields:
            logger.warning("⚠️ ATENÇÃO: Nenhuma coluna de CAT. (aba PRODUTO) foi mapeada!")
            logger.error("🚨 Colunas disponíveis que contêm 'cat':")
            for col in df.columns:
                if "cat" in col.lower():
                    logger.error(f"    - '{col}'")

        if "grupo" not in mapped_fields:
            logger.warning("⚠️ ATENÇÃO: Nenhuma coluna de GRUPO foi mapeada!")

        if not rename_dict:
            available_cols = ", ".join(df.columns[:10])
            raise ExcelProcessingError(
                f"Nenhuma coluna reconhecida encontrada. "
                f"Colunas disponíveis: {available_cols}..."
            )

        logger.info(f"📋 Total de colunas mapeadas: {len(rename_dict)}")

        # ✅ DEBUG: Verificar dados após mapeamento (CORRIGIDO)
        df_mapped = df.rename(rename_dict)

        if "categoria" in df_mapped.columns:
            logger.info("🔍 === VERIFICANDO DADOS DA CATEGORIA (LOJA WEB) APÓS MAPEAMENTO ===")
            categoria_sample = df_mapped.select("categoria").head(10).to_series().to_list()
            logger.info(f"  Primeiras 10 categorias: {categoria_sample}")

        if "cat" in df_mapped.columns:
            logger.info("🔍 === VERIFICANDO DADOS DA CAT. (aba PRODUTO) APÓS MAPEAMENTO ===")
            cat_sample = df_mapped.select("cat").head(10).to_series().to_list()
            logger.info(f"  Primeiras 10 Cat.: {cat_sample}")

        if "grupo" in df_mapped.columns:
            logger.info("🔍 === VERIFICANDO DADOS DO GRUPO APÓS MAPEAMENTO ===")
            grupo_sample = df_mapped.select("grupo").head(10).to_series().to_list()
            logger.info(f"  Primeiros 10 grupos: {grupo_sample}")

        return df_mapped

    def _is_ean_variacao_column(self, col_name: str) -> bool:
        """Verifica se a coluna é EAN e Variação"""
        if not col_name:
            return False

        col_normalized = self._normalize_column_name(col_name)
        return "ean" in col_normalized and "variacao" in col_normalized

    def _is_tipo_produto_column(self, col_name: str) -> bool:
        """Verifica se a coluna é Tipo de Produto"""
        if not col_name:
            return False

        col_normalized = self._normalize_column_name(col_name)
        return "tipo" in col_normalized and "produto" in col_normalized

    def _is_cor_produto_column(self, col_name: str) -> bool:
        """Verifica se a coluna é Cor do Produto"""
        if not col_name:
            return False

        col_normalized = self._normalize_column_name(col_name)
        return "cor" in col_normalized and "produto" in col_normalized

    def _is_cor_simples_column(self, col_name: str) -> bool:
        """Verifica se a coluna é apenas 'COR' ou 'Cor'"""
        if not col_name:
            return False

        col_clean = col_name.strip().lower()
        return col_clean == "cor"

    def _is_cor_tecido_column(self, col_name: str) -> bool:
        """Verifica se a coluna é Cor do tecido (FALLBACK)"""
        if not col_name:
            return False

        col_normalized = self._normalize_column_name(col_name)
        return "cor" in col_normalized and "tecido" in col_normalized

    def _is_titulo_compra_column(self, col_name: str) -> bool:
        """Verifica se a coluna é Título para Compra"""
        if not col_name:
            return False

        col_normalized = self._normalize_column_name(col_name)
        return "titulo" in col_normalized and "compra" in col_normalized

    def _is_complemento_sem_cod_column(self, col_name: str) -> bool:
        """Verifica se a coluna é Complemento do Produto (sem o cod e marca)"""
        if not col_name:
            return False

        col_normalized = self._normalize_column_name(col_name)
        return ("complemento" in col_normalized and
                "produto" in col_normalized and
                "sem" in col_normalized)

    def _is_anuncio_column(self, col_name: str) -> bool:
        """Verifica se a coluna é de anúncio"""
        if not col_name:
            return False

        col_normalized = self._normalize_column_name(col_name)

        # Variações possíveis de "anúncio"
        anuncio_variants = [
            "anuncio", "anúncio", "titulo", "tituloproduto",
            "nomeanuncio", "descricaoanuncio", "nomecomercial"
        ]

        # Verifica se alguma variante está presente
        for variant in anuncio_variants:
            if variant in col_normalized:
                return True

        # Verifica também o nome original
        col_original = col_name.lower().strip()
        return any(variant in col_original for variant in ["anúncio", "anuncio"])

    def _is_categoria_column(self, col_name: str) -> bool:
        """Verifica se a coluna é de categoria (LOJA WEB)"""
        if not col_name:
            return False

        col_normalized = self._normalize_column_name(col_name)
        col_lower = col_name.lower().strip()

        # Variações possíveis para categoria (LOJA WEB)
        categoria_variants = [
            "categoria", "categorias", "idcategoria", "codcategoria",
            "codigocategoria"
        ]

        # Verifica se alguma variante está presente
        for variant in categoria_variants:
            if variant in col_normalized or variant in col_lower:
                return True

        return False

    def _is_cat_column(self, col_name: str) -> bool:
        """Verifica se a coluna é Cat. (aba PRODUTO)"""
        if not col_name:
            return False

        col_normalized = self._normalize_column_name(col_name)
        col_lower = col_name.lower().strip()

        # Variações possíveis para Cat. (aba PRODUTO)
        cat_variants = ["cat", "cat."]

        # Verifica se alguma variante está presente
        for variant in cat_variants:
            if variant in col_normalized or variant in col_lower:
                # Evita confundir com "categoria"
                if "categoria" not in col_lower:
                    return True

        return False

    def _is_grupo_column(self, col_name: str) -> bool:
        """Verifica se a coluna é Grupo"""
        if not col_name:
            return False

        col_normalized = self._normalize_column_name(col_name)
        col_lower = col_name.lower().strip()

        # Variações possíveis
        grupo_variants = ["grupo", "grupos"]

        # Verifica se alguma variante está presente
        for variant in grupo_variants:
            if variant in col_normalized or variant in col_lower:
                return True

        return False

    def _clean_data(self, df: pl.DataFrame) -> pl.DataFrame:
        """Limpa e normaliza dados - VERSÃO COM DEBUG INTENSIVO"""
        try:
            logger.info("🔍 === INICIANDO _clean_data ===")
            logger.info(f"DataFrame shape: {df.shape}")
            logger.info(f"Colunas: {df.columns}")

            # Verificar tipos das colunas
            logger.info("🔍 === TIPOS DAS COLUNAS ===")
            for col in df.columns:
                dtype = df[col].dtype
                logger.info(f"  {col}: {dtype}")

            # Lista de colunas que devem existir
            required_cols = ["ean"]

            # Verifica se colunas obrigatórias existem
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                raise ExcelProcessingError(f"Colunas obrigatórias não encontradas: {missing_cols}")

            logger.info(f"Colunas disponíveis após mapeamento: {df.columns}")

            # ✅ ESTRATÉGIA MAIS SEGURA: PROCESSAR UMA COLUNA POR VEZ
            logger.info("🔍 === PROCESSANDO COLUNAS UMA POR VEZ ===")

            # ✅ 1. TRATAR EAN PRIMEIRO
            if "ean" in df.columns:
                try:
                    logger.info(f"🧹 Processando coluna EAN...")
                    logger.info(f"  Tipo atual: {df['ean'].dtype}")

                    # Verificar alguns valores
                    sample_ean = df.select("ean").head(3).to_series().to_list()
                    logger.info(f"  Amostra EAN: {sample_ean}")

                    df = df.with_columns([
                        pl.col("ean")
                        .cast(pl.Utf8, strict=False)
                        .fill_null("")
                        .str.replace(r"\.0+$", "", literal=False)
                        .str.strip_chars()
                        .alias("ean")
                    ])
                    logger.info("✅ EAN processado com sucesso")

                except Exception as ean_error:
                    logger.error(f"❌ Erro ao processar EAN: {ean_error}")
                    raise

            # ✅ 2. TRATAR CATEGORIA SEGUNDO (LOJA WEB)
            if "categoria" in df.columns:
                try:
                    logger.info(f"🧹 Processando coluna CATEGORIA (LOJA WEB)...")
                    logger.info(f"  Tipo atual: {df['categoria'].dtype}")

                    # Verificar alguns valores
                    sample_cat = df.select("categoria").head(3).to_series().to_list()
                    logger.info(f"  Amostra CATEGORIA: {sample_cat}")

                    df = df.with_columns([
                        pl.col("categoria")
                        .cast(pl.Utf8, strict=False)
                        .fill_null("")
                        .str.replace(r"\.0+$", "", literal=False)
                        .str.strip_chars()
                        .alias("categoria")
                    ])
                    logger.info("✅ CATEGORIA (LOJA WEB) processada com sucesso")

                except Exception as cat_error:
                    logger.error(f"❌ Erro ao processar CATEGORIA: {cat_error}")
                    raise

            # ✅ 3. TRATAR CAT. (aba PRODUTO)
            if "cat" in df.columns:
                try:
                    logger.info(f"🧹 Processando coluna CAT. (aba PRODUTO)...")
                    logger.info(f"  Tipo atual: {df['cat'].dtype}")

                    # Verificar alguns valores
                    sample_cat = df.select("cat").head(3).to_series().to_list()
                    logger.info(f"  Amostra CAT.: {sample_cat}")

                    df = df.with_columns([
                        pl.col("cat")
                        .cast(pl.Utf8, strict=False)
                        .fill_null("")
                        .str.replace(r"\.0+$", "", literal=False)
                        .str.strip_chars()
                        .alias("cat")
                    ])
                    logger.info("✅ CAT. (aba PRODUTO) processada com sucesso")

                except Exception as cat_error:
                    logger.error(f"❌ Erro ao processar CAT.: {cat_error}")
                    raise

            # ✅ 4. PROCESSAR OUTRAS COLUNAS STRING UMA POR VEZ
            string_columns = [
                "ean_variacao", "cod_fornecedor", "complemento_titulo",
                "anuncio", "titulo_compra", "grupo", "cor", "tipo_anuncio", "tipo_produto",
                "descricao_html", "ncm", "complemento_produto", "produto"  # produto opcional
            ]

            for col in string_columns:
                if col in df.columns:
                    try:
                        logger.info(f"🧹 Processando coluna STRING: {col}")
                        logger.info(f"  Tipo atual: {df[col].dtype}")

                        df = df.with_columns([
                            pl.col(col)
                            .cast(pl.Utf8, strict=False)
                            .fill_null("")
                            .str.strip_chars()
                            .alias(col)
                        ])
                        logger.info(f"✅ {col} processado com sucesso")

                    except Exception as col_error:
                        logger.error(f"❌ Erro ao processar {col}: {col_error}")
                        # Continua sem essa coluna
                        continue

            # ✅ 5. PROCESSAR COLUNAS NUMÉRICAS (SE EXISTIREM)
            int_columns = ["volumes", "qtde_volume", "prazo"]
            for col in int_columns:
                if col in df.columns:
                    try:
                        logger.info(f"🔢 Processando coluna INT: {col}")
                        logger.info(f"  Tipo atual: {df[col].dtype}")

                        df = df.with_columns([
                            pl.col(col)
                            .fill_null(1 if col in ["volumes", "qtde_volume"] else 0)
                            .cast(pl.Int32, strict=False)
                            .alias(col)
                        ])
                        logger.info(f"✅ {col} processado com sucesso")

                    except Exception as col_error:
                        logger.error(f"❌ Erro ao processar {col}: {col_error}")
                        continue

            float_columns = ["peso_bruto", "peso_liquido", "largura", "altura", "comprimento"]
            for col in float_columns:
                if col in df.columns:
                    try:
                        logger.info(f"🔢 Processando coluna FLOAT: {col}")
                        logger.info(f"  Tipo atual: {df[col].dtype}")

                        df = df.with_columns([
                            pl.col(col)
                            .fill_null(0.0)
                            .cast(pl.Float64, strict=False)
                            .alias(col)
                        ])
                        logger.info(f"✅ {col} processado com sucesso")

                    except Exception as col_error:
                        logger.error(f"❌ Erro ao processar {col}: {col_error}")
                        continue

            logger.info("✅ Todas as colunas processadas")

            # ✅ DEBUG FINAL: Verificar dados após limpeza
            try:
                if "categoria" in df.columns:
                    logger.info("�� === DADOS DA CATEGORIA (LOJA WEB) APÓS LIMPEZA ===")
                    categoria_sample = df.select("categoria").head(10).to_series().to_list()
                    logger.info(f"  Amostra de categorias: {categoria_sample}")
                    logger.info(f"  Tipo final: {df['categoria'].dtype}")

                if "cat" in df.columns:
                    logger.info("🔍 === DADOS DA CAT. (aba PRODUTO) APÓS LIMPEZA ===")
                    cat_sample = df.select("cat").head(10).to_series().to_list()
                    logger.info(f"  Amostra de Cat.: {cat_sample}")
                    logger.info(f"  Tipo final: {df['cat'].dtype}")

                if "grupo" in df.columns:
                    logger.info("�� === DADOS DO GRUPO APÓS LIMPEZA ===")
                    grupo_sample = df.select("grupo").head(10).to_series().to_list()
                    logger.info(f"  Amostra de grupos: {grupo_sample}")
                    logger.info(f"  Tipo final: {df['grupo'].dtype}")

            except Exception as debug_error:
                logger.warning(f"⚠️ Erro no debug final: {debug_error}")

            # ✅ FILTRAR LINHAS VAZIAS (SEM COMPARAÇÕES COMPLEXAS)
            try:
                logger.info("🔍 === FILTRANDO LINHAS VAZIAS ===")
                original_count = df.height

                # Filtro mais simples possível
                df = df.filter(
                    pl.col("ean").is_not_null() &
                    (pl.col("ean") != "") &
                    (pl.col("ean") != "nan")
                )

                final_count = df.height
                logger.info(f"✅ Filtro aplicado: {final_count}/{original_count} linhas mantidas")

            except Exception as filter_error:
                logger.error(f"❌ Erro ao filtrar: {filter_error}")
                logger.warning("⚠️ Continuando sem filtrar")

            logger.info("✅ _clean_data concluído com sucesso")
            return df

        except Exception as e:
            logger.error(f"❌ Erro na limpeza de dados: {e}")
            import traceback
            logger.error(f"Stack trace: {traceback.format_exc()}")
            raise ExcelProcessingError(f"Falha na limpeza de dados: {e}")

    def _clean_row_data(self, row: dict) -> dict:
        """Limpa dados de uma linha específica"""
        clean_row = {}

        for key, value in row.items():
            if value is None:
                # Define valores padrão baseado no tipo esperado
                if key in ["volumes", "prazo"]:
                    clean_row[key] = 1 if key == "volumes" else 0
                elif key == "qtde_volume":
                    clean_row[key] = None  # ✅ Deixa None para detectar na descrição
                else:
                    clean_row[key] = ""
            elif isinstance(value, str):
                clean_row[key] = value.strip()
            else:
                clean_row[key] = value

        return clean_row

    def read_categories(self, file_path: Path) -> Dict[str, Tuple[str, str, str]]:
        """Lê mapeamento de categorias"""
        try:
            logger.info(f"Carregando categorias de: {file_path}")

            if not file_path.exists():
                logger.warning(f"Arquivo de categorias não encontrado: {file_path}")
                return {}

            # ✅ USAR PANDAS PRIMEIRO, DEPOIS POLARS
            try:
                import pandas as pd
                df_pandas = pd.read_excel(file_path, engine='openpyxl')
                df = pl.from_pandas(df_pandas)
            except Exception:
                df = pl.read_excel(file_path, engine="openpyxl")

            # Mapeia categorias
            category_map = {}
            for row in df.iter_rows(named=True):
                principal = row.get("CATEGORIA PRINCIPAL", "")
                nivel1 = row.get("NIVEL ADICIONAL 1", "")
                nivel2 = row.get("NIVEL ADICIONAL 2", "")

                if principal and str(principal).strip():
                    key = str(principal).strip()
                    category_map[key] = (
                        str(principal).strip(),
                        str(nivel1).strip() if nivel1 else "",
                        str(nivel2).strip() if nivel2 else ""
                    )

            logger.info(f"📂 {len(category_map)} categorias carregadas")
            return category_map

        except Exception as e:
            logger.warning(f"⚠️ Erro ao carregar categorias: {e}")
            return {}

    def get_sheet_names(self, file_path: Path) -> List[str]:
        """Retorna lista de nomes das abas do arquivo Excel"""
        try:
            if not file_path.exists():
                logger.warning(f"Arquivo não encontrado: {file_path}")
                return []

            logger.info(f"📋 Lendo abas do arquivo: {file_path}")

            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            logger.info(f"✅ {len(sheet_names)} abas encontradas: {sheet_names}")
            return sheet_names

        except Exception as e:
            logger.error(f"❌ Erro ao ler abas do arquivo: {e}")
            return []