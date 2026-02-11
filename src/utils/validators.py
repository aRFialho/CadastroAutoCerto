import re
from pathlib import Path
from typing import List
from ..core.exceptions import ValidationError


def validate_ean(ean: str) -> bool:
    """Valida código EAN"""
    if not ean:
        return False

    # Remove espaços e caracteres especiais
    ean_clean = re.sub(r'[^0-9]', '', str(ean))

    # EAN deve ter 8, 12, 13 ou 14 dígitos
    if len(ean_clean) not in [8, 12, 13, 14]:
        return False

    return True


def validate_excel_file(file_path: Path) -> bool:
    """Valida se é um arquivo Excel válido"""
    if not file_path.exists():
        raise ValidationError(f"Arquivo não encontrado: {file_path}")

    if file_path.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValidationError(f"Arquivo deve ser Excel (.xlsx ou .xls): {file_path}")

    return True


def validate_email(email: str) -> bool:
    """Valida formato de e-mail"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_required_fields(data: dict, required_fields: List[str]) -> List[str]:
    """Valida campos obrigatórios"""
    missing_fields = []

    for field in required_fields:
        if field not in data or not data[field] or str(data[field]).strip() == "":
            missing_fields.append(field)

    return missing_fields


def clean_numeric_value(value, default=0.0):
    """Limpa e converte valor numérico"""
    if value is None or value == "":
        return default

    try:
        # Remove espaços e troca vírgula por ponto
        clean_value = str(value).strip().replace(',', '.')
        return float(clean_value)
    except (ValueError, TypeError):
        return default


def extract_quantity_from_title(title: str) -> int:
    """Extrai quantidade do título (ex: '2x Produto' -> 2)"""
    if not title:
        return 1

    # Procura padrões como "2x", "x3", "2 x", etc.
    patterns = [
        r'(\d+)\s*[xX]',  # 2x, 2 x
        r'[xX]\s*(\d+)',  # x2, x 2
        r'(\d+)\s*un',  # 2un, 2 un
        r'(\d+)\s*pç',  # 2pç, 2 pç
    ]

    for pattern in patterns:
        match = re.search(pattern, title, re.IGNORECASE)
        if match:
            try:
                return int(match.group(1))
            except ValueError:
                continue

    return 1


def normalize_category(category: str) -> str:
    """Normaliza nome de categoria"""
    if not category:
        return ""

    # Remove espaços extras e normaliza
    normalized = re.sub(r'\s+', ' ', str(category).strip())

    # Capitaliza primeira letra de cada palavra
    normalized = normalized.title()

    return normalized


def validate_sheet_exists(file_path: Path, sheet_name: str) -> bool:
    """Valida se aba existe no arquivo Excel"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        return sheet_name in wb.sheetnames
    except Exception:
        return False


def get_available_sheets(file_path: Path) -> List[str]:
    """Retorna lista de abas disponíveis"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        return wb.sheetnames
    except Exception:
        return []
