from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from .athos_models import AthosAction, RuleName, ORDERED_RULES


# =========================
#  Gerador Excel (Template)
# =========================

REQUIRED_SHEET_NAME = "PRODUTO"

# Headers exatos (vamos fazer match por normalização)
HDR_CODBARRA = "Código de Barras"
HDR_GRUPO3 = "GRUPO3"
HDR_ESTOQUE_SEG = "Estoque de Segurança"
HDR_PROD_INATIVO = "Produto Inativo"
HDR_DIAS_ENTREGA = "Dias para Entrega"
HDR_SITE_DISP = "Site Disponibilidade"


def _norm_header(s: str) -> str:
    return str(s or "").strip().lower()


@dataclass
class GeneratedFiles:
    # regra -> caminho do arquivo gerado
    files_by_rule: Dict[RuleName, Path]


def generate_rule_files(
    template_path: str | Path,
    out_dir: str | Path,
    actions_by_rule: Dict[RuleName, List[AthosAction]],
    date_tag: str,
) -> GeneratedFiles:
    """
    Gera 5 arquivos (um por regra), preenchendo somente a aba PRODUTO.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Template não encontrado: {template_path}")

    files_by_rule: Dict[RuleName, Path] = {}

    for idx, rule in enumerate(ORDERED_RULES, start=1):
        actions = actions_by_rule.get(rule, [])
        # Sempre gera o arquivo, mesmo vazio? (recomendado: sim, para manter padrão)
        out_name = f"{idx:02d}_{rule.value.replace(' ', '_')}_{date_tag}.xlsx"
        out_path = out_dir / out_name
        _write_one_file(template_path, out_path, actions)
        files_by_rule[rule] = out_path

    return GeneratedFiles(files_by_rule=files_by_rule)


def _write_one_file(template_path: Path, out_path: Path, actions: List[AthosAction]) -> None:
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise RuntimeError("openpyxl não está instalado. Instale para gerar .xlsx") from e

    wb = openpyxl.load_workbook(template_path)
    if REQUIRED_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Template não possui aba '{REQUIRED_SHEET_NAME}'. Abas: {wb.sheetnames}")

    ws = wb[REQUIRED_SHEET_NAME]

    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError("Não consegui localizar a linha de cabeçalho na aba PRODUTO (procurei por 'Código de Barras').")

    col_map = _build_col_map(ws, header_row)

    # limpa conteúdo antigo (abaixo do cabeçalho) nas colunas relevantes
    _clear_previous(ws, header_row + 1, col_map)

    # escreve linhas a partir de header_row+1
    row_ptr = header_row + 1
    for a in actions:
        # Código de Barras obrigatório
        ws.cell(row=row_ptr, column=col_map[HDR_CODBARRA]).value = a.codbarra

        if HDR_GRUPO3 in col_map and a.grupo3 is not None:
            ws.cell(row=row_ptr, column=col_map[HDR_GRUPO3]).value = a.grupo3

        if HDR_ESTOQUE_SEG in col_map and a.estoque_seguranca is not None:
            ws.cell(row=row_ptr, column=col_map[HDR_ESTOQUE_SEG]).value = int(a.estoque_seguranca)

        if HDR_PROD_INATIVO in col_map and a.produto_inativo is not None:
            ws.cell(row=row_ptr, column=col_map[HDR_PROD_INATIVO]).value = a.produto_inativo

        if HDR_DIAS_ENTREGA in col_map and a.dias_entrega is not None:
            ws.cell(row=row_ptr, column=col_map[HDR_DIAS_ENTREGA]).value = int(a.dias_entrega)

        if HDR_SITE_DISP in col_map and a.site_disponibilidade is not None:
            ws.cell(row=row_ptr, column=col_map[HDR_SITE_DISP]).value = a.site_disponibilidade

        row_ptr += 1

    wb.save(out_path)


def _find_header_row(ws) -> Optional[int]:
    """
    Procura a linha de cabeçalho olhando onde aparece "Código de Barras".
    """
    try:
        max_r = ws.max_row
        max_c = ws.max_column
    except Exception:
        return None

    target = _norm_header(HDR_CODBARRA)

    for r in range(1, min(max_r, 50) + 1):  # cabeçalho deve estar no topo
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if _norm_header(v) == target:
                return r
    return None


def _build_col_map(ws, header_row: int) -> Dict[str, int]:
    """
    Mapeia nomes de colunas -> índice (1-based) baseado nos headers.
    """
    desired = {
        _norm_header(HDR_CODBARRA): HDR_CODBARRA,
        _norm_header(HDR_GRUPO3): HDR_GRUPO3,
        _norm_header(HDR_ESTOQUE_SEG): HDR_ESTOQUE_SEG,
        _norm_header(HDR_PROD_INATIVO): HDR_PROD_INATIVO,
        _norm_header(HDR_DIAS_ENTREGA): HDR_DIAS_ENTREGA,
        _norm_header(HDR_SITE_DISP): HDR_SITE_DISP,
    }

    col_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        key = _norm_header(v)
        if key in desired:
            col_map[desired[key]] = c

    # obrigatório: Código de Barras
    if HDR_CODBARRA not in col_map:
        raise ValueError("Template não tem coluna 'Código de Barras' na aba PRODUTO.")

    return col_map


def _clear_previous(ws, start_row: int, col_map: Dict[str, int]) -> None:
    """
    Limpa conteúdo antigo nas colunas usadas, do start_row até a última linha.
    """
    last = ws.max_row
    cols = list(col_map.values())
    for r in range(start_row, last + 1):
        for c in cols:
            ws.cell(row=r, column=c).value = None
