from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

from .athos_models import AthosReportLine


@dataclass
class ReportFile:
    path: Path
    total_lines: int


def write_report_xlsx(
    report_lines: List[AthosReportLine],
    out_dir: str | Path,
    date_tag: str,
    filename_prefix: str = "RELATORIO",
) -> ReportFile:
    """
    Gera um XLSX com 1 aba (RELATORIO) contendo:
    Planilha | Código de Barras | Tipo | Marca | Grupo3 | Ação
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    out_path = out_dir / f"{filename_prefix}_{date_tag}.xlsx"

    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise RuntimeError("openpyxl não está instalado. Instale para gerar relatório .xlsx") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RELATORIO"

    headers = ["PLANILHA", "COD_BARRA", "TIPO", "MARCA", "GRUPO3", "ACAO"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = h

    r = 2
    for line in report_lines:
        ws.cell(row=r, column=1).value = line.planilha
        ws.cell(row=r, column=2).value = line.codbarra
        ws.cell(row=r, column=3).value = line.tipo
        ws.cell(row=r, column=4).value = line.marca
        ws.cell(row=r, column=5).value = line.grupo3
        ws.cell(row=r, column=6).value = line.acao
        r += 1

    # Congelar header
    ws.freeze_panes = "A2"

    # Ajuste simples de largura
    col_widths = [18, 20, 8, 22, 20, 55]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(out_path)

    return ReportFile(path=out_path, total_lines=len(report_lines))
