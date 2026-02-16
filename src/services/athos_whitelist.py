from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional, Set, Tuple


# =========================
#  Whitelist (Imediatos)
# =========================

CANDIDATE_COLS = [
    "cod barra",
    "codbarras",
    "codbarra",
    "ean",
    "código de barras",
    "codigo de barras",
]


@dataclass
class WhitelistLoadResult:
    path: Path
    eans: Set[str]
    total_rows_seen: int
    valid_eans: int
    duplicates_ignored: int
    invalid_ignored: int
    detected_column: Optional[str] = None


def normalize_ean(value) -> Optional[str]:
    """
    Normaliza um EAN/código de barras:
    - converte para string
    - remove espaços
    - remove '.0' típico do Excel quando vem como float
    - remove tudo que não for dígito
    - retorna None se ficar vazio
    """
    if value is None:
        return None

    s = str(value).strip()

    if not s or s.lower() in ("nan", "none"):
        return None

    # remove .0 típico de excel float
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]

    # manter apenas dígitos
    s = re.sub(r"\D+", "", s)

    if not s:
        return None
    return s


def _detect_column(columns: Iterable[str]) -> Optional[str]:
    """
    Detecta a coluna de EAN pela lista de candidatos.
    Retorna o nome original da coluna (como está no arquivo) ou None.
    """
    cols = list(columns)
    normalized = {str(c).strip().lower(): c for c in cols}

    for cand in CANDIDATE_COLS:
        if cand in normalized:
            return str(normalized[cand])

    # fallback: qualquer coluna que contenha "barra" ou "ean"
    for k, original in normalized.items():
        if "barra" in k or "ean" in k:
            return str(original)

    return None


def copy_whitelist_to_outputs(selected_file: str | Path) -> Path:
    """
    Copia o arquivo escolhido pelo usuário para:
      outputs/imediatos/whitelist_imediatos.<ext>

    Retorna o caminho final salvo.
    """
    src = Path(selected_file)
    if not src.exists():
        raise FileNotFoundError(f"Arquivo de whitelist não encontrado: {src}")

    dest_dir = Path("outputs") / "imediatos"
    dest_dir.mkdir(parents=True, exist_ok=True)

    ext = src.suffix.lower() or ".dat"
    dest = dest_dir / f"whitelist_imediatos{ext}"

    shutil.copy2(src, dest)
    return dest


def load_whitelist(file_path: str | Path) -> WhitelistLoadResult:
    """
    Carrega whitelist de imediatos a partir de:
    - .xls / .xlsx (Excel)
    - .csv
    - .txt (1 por linha)

    Retorna set de EANs normalizados + estatísticas.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Whitelist não encontrada: {path}")

    suffix = path.suffix.lower()

    if suffix in (".txt",):
        return _load_from_txt(path)

    if suffix in (".csv",):
        return _load_from_csv(path)

    if suffix in (".xlsx", ".xls"):
        return _load_from_excel(path)

    # fallback: tenta txt
    return _load_from_txt(path)


def _load_from_txt(path: Path) -> WhitelistLoadResult:
    eans: Set[str] = set()
    duplicates = 0
    invalid = 0
    total = 0

    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            total += 1
            ean = normalize_ean(line)
            if not ean:
                invalid += 1
                continue
            if ean in eans:
                duplicates += 1
            eans.add(ean)

    return WhitelistLoadResult(
        path=path,
        eans=eans,
        total_rows_seen=total,
        valid_eans=len(eans),
        duplicates_ignored=duplicates,
        invalid_ignored=invalid,
        detected_column=None,
    )


def _load_from_csv(path: Path) -> WhitelistLoadResult:
    # tenta pandas primeiro
    try:
        import pandas as pd  # type: ignore
    except Exception:
        pd = None  # type: ignore

    if pd is not None:
        df = pd.read_csv(path, dtype=str, sep=None, engine="python")
        col = _detect_column(df.columns)
        if col is None:
            # se só tiver 1 coluna, usa ela
            if len(df.columns) == 1:
                col = str(df.columns[0])
            else:
                raise ValueError(f"Não consegui detectar a coluna de EAN no CSV: {list(df.columns)}")

        series = df[col].tolist()
        return _build_result_from_values(path, series, detected_column=col)

    # fallback sem pandas: lê primeira coluna por ';' ou ','
    eans: Set[str] = set()
    duplicates = 0
    invalid = 0
    total = 0

    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            total += 1
            parts = re.split(r"[;,]", line)
            ean = normalize_ean(parts[0])
            if not ean:
                invalid += 1
                continue
            if ean in eans:
                duplicates += 1
            eans.add(ean)

    return WhitelistLoadResult(
        path=path,
        eans=eans,
        total_rows_seen=total,
        valid_eans=len(eans),
        duplicates_ignored=duplicates,
        invalid_ignored=invalid,
        detected_column=None,
    )


def _load_from_excel(path: Path) -> WhitelistLoadResult:
    """
    Suporta .xlsx via openpyxl e tenta .xls via pandas/xlrd se disponível.
    """
    suffix = path.suffix.lower()

    # 1) tenta pandas (lê xls/xlsx fácil)
    try:
        import pandas as pd  # type: ignore

        df = pd.read_excel(path, dtype=str)
        col = _detect_column(df.columns)
        if col is None:
            if len(df.columns) == 1:
                col = str(df.columns[0])
            else:
                raise ValueError(f"Não consegui detectar a coluna de EAN no Excel: {list(df.columns)}")

        return _build_result_from_values(path, df[col].tolist(), detected_column=col)
    except Exception:
        pass

    # 2) fallback: openpyxl para .xlsx
    if suffix == ".xlsx":
        try:
            import openpyxl  # type: ignore
        except Exception as e:
            raise RuntimeError(f"openpyxl não disponível para ler .xlsx: {e}")

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        # detect header row = 1
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers_str = ["" if h is None else str(h) for h in headers]
        col_name = _detect_column(headers_str)
        if col_name is None:
            # se só 1 coluna, assume A
            if ws.max_column == 1:
                col_idx = 1
                detected = headers_str[0] if headers_str else None
            else:
                raise ValueError(f"Não consegui detectar a coluna de EAN no Excel (.xlsx). Headers: {headers_str}")
        else:
            col_idx = headers_str.index(col_name) + 1
            detected = col_name

        values = []
        # começa na linha 2
        for r in range(2, ws.max_row + 1):
            values.append(ws.cell(row=r, column=col_idx).value)

        return _build_result_from_values(path, values, detected_column=detected)

    # 3) se .xls e não conseguiu via pandas, falha com msg clara
    raise RuntimeError(
        "Não consegui ler arquivo .xls. Sugestões: "
        "1) instale 'xlrd' (para .xls) ou "
        "2) exporte a whitelist como .xlsx/.csv/.txt."
    )


def _build_result_from_values(
    path: Path,
    values: Iterable,
    detected_column: Optional[str] = None,
) -> WhitelistLoadResult:
    eans: Set[str] = set()
    duplicates = 0
    invalid = 0
    total = 0

    for v in values:
        total += 1
        ean = normalize_ean(v)
        if not ean:
            invalid += 1
            continue
        if ean in eans:
            duplicates += 1
        eans.add(ean)

    return WhitelistLoadResult(
        path=path,
        eans=eans,
        total_rows_seen=total,
        valid_eans=len(eans),
        duplicates_ignored=duplicates,
        invalid_ignored=invalid,
        detected_column=detected_column,
    )
