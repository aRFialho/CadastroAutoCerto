"""
AthosExcelWriter
- Abre o template (já copiado) e preenche apenas a aba "PRODUTOS"
- Escreve relatório consolidado em .txt

Foco:
✅ preservar layout do template (só escreve valores)
✅ robusto com headers (acha coluna por nome)
✅ limpa linhas anteriores (dados) antes de escrever novos
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from ..utils.logger import get_logger

logger = get_logger("athos_excel_writer")


@dataclass
class SheetWriteResult:
    rows_written: int
    warnings: List[str]


class AthosExcelWriter:
    def __init__(self) -> None:
        pass

    # =========================
    # Public API
    # =========================
    def write_rule_workbook(
        self,
        workbook_path: Path,
        rows: List[Dict[str, Any]],
        sheet_name: str = "PRODUTOS",
        clear_existing_data: bool = True,
    ) -> SheetWriteResult:
        """
        Escreve linhas na aba PRODUTOS do workbook.
        - rows: lista de dicts {header_name: value}
        - SEMPRE tenta preencher EAN/Cód Barra se existir header correspondente
        """
        workbook_path = Path(workbook_path)
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook não encontrado: {workbook_path}")

        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise RuntimeError(f"openpyxl não disponível: {e}")

        wb = load_workbook(workbook_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não existe no template: {workbook_path.name}")

        ws = wb[sheet_name]

        header_row_idx, header_map = self._detect_header(ws)
        if header_row_idx is None or not header_map:
            raise ValueError(
                f"Não consegui detectar cabeçalho na aba '{sheet_name}'. "
                f"Verifique se o template possui headers na primeira linha."
            )

        warnings: List[str] = []

        # Limpar linhas existentes (mantendo cabeçalho)
        if clear_existing_data:
            self._clear_data_below_header(ws, header_row_idx)

        # Escrever linhas
        start_row = header_row_idx + 1
        current_row = start_row

        # normaliza keys de rows para bater com header_map
        for item in rows:
            if not item:
                continue

            # Preenche colunas por header (case-insensitive)
            for k, v in item.items():
                col_idx = self._find_col(header_map, k)
                if col_idx is None:
                    # ignorar colunas que não existem no template
                    continue
                ws.cell(row=current_row, column=col_idx, value=v)

            current_row += 1

        wb.save(workbook_path)

        written = max(0, current_row - start_row)
        logger.info(f"[AthosExcelWriter] {workbook_path.name} -> {written} linhas gravadas em '{sheet_name}'")
        return SheetWriteResult(rows_written=written, warnings=warnings)

    def write_report_txt(self, report_path: Path, lines: List[str]) -> None:
        report_path = Path(report_path)
        report_path.write_text("\n".join(lines), encoding="utf-8")
        logger.info(f"[AthosExcelWriter] Relatório gravado: {report_path.name}")

    # =========================
    # Internals
    # =========================
    def _normalize(self, s: Any) -> str:
        if s is None:
            return ""
        return str(s).strip().lower()

    def _detect_header(self, ws) -> tuple[Optional[int], Dict[str, int]]:
        """
        Detecta a linha de cabeçalho:
        - procura a primeira linha que tenha pelo menos 2 células com texto
        Retorna (row_index, {header_normalizado: col_index})
        """
        max_scan = min(ws.max_row or 200, 200)

        best_row = None
        best_count = 0
        best_map: Dict[str, int] = {}

        for r in range(1, max_scan + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, (ws.max_column or 50) + 1)]
            texts = [self._normalize(v) for v in row_vals if v is not None and str(v).strip() != ""]
            # heurística simples: cabeçalho costuma ter várias colunas preenchidas
            if len(texts) >= 2 and len(texts) > best_count:
                header_map: Dict[str, int] = {}
                for c in range(1, (ws.max_column or 50) + 1):
                    hv = ws.cell(row=r, column=c).value
                    hs = self._normalize(hv)
                    if hs:
                        header_map[hs] = c
                best_row = r
                best_count = len(texts)
                best_map = header_map

        return best_row, best_map

    def _clear_data_below_header(self, ws, header_row_idx: int) -> None:
        """
        Limpa valores das linhas abaixo do header (não mexe no estilo).
        """
        if ws.max_row is None or ws.max_row <= header_row_idx:
            return

        # define limite de colunas com base no header
        max_col = ws.max_column or 50

        for r in range(header_row_idx + 1, (ws.max_row or header_row_idx) + 1):
            empty_row = True
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value not in (None, ""):
                    empty_row = False
                    break
            # se já estiver vazio, pode continuar (mas não para, porque pode ter buracos)
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c, value=None)

    def _find_col(self, header_map: Dict[str, int], key: str) -> Optional[int]:
        """
        Encontra coluna pelo nome do header (case-insensitive e tolerante).
        """
        nk = self._normalize(key)
        if not nk:
            return None

        # match direto
        if nk in header_map:
            return header_map[nk]

        # match tolerante: remove espaços múltiplos
        nk2 = " ".join(nk.split())
        if nk2 in header_map:
            return header_map[nk2]

        # match por "contém" (ex.: "cód barra" vs "cod barra")
        for hk, idx in header_map.items():
            if nk2 == hk:
                return idx
            if nk2 and hk and (nk2 in hk or hk in nk2):
                return idx

        return None
