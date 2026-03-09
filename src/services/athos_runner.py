# src/services/athos_runner.py
"""
Service: AthosRunner

Responsável por:
- Ler o export do SQL (Excel)
- Ler whitelist (PRODUTOS.xls/.xlsx) -> lista de EANs "imediatos"
- Abrir template Athos e gerar 5 planilhas (cada uma com aba 'PRODUTO(S)')
- Aplicar regras e preencher SOMENTE a aba 'PRODUTO(S)' do template
- Gerar relatório consolidado

Saídas geradas (na ordem):
01_FORA_DE_LINHA.xlsx
02_ESTOQUE_COMPARTILHADO.xlsx
03_ENVIO_IMEDIATO.xlsx
04_SEM_GRUPO.xlsx
05_OUTLET.xlsx
RELATORIO_CONSOLIDADO.xlsx
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from datetime import datetime
from typing import Callable, Optional, Any, Dict, List, Iterable
import shutil
from collections import Counter

from ..utils.logger import get_logger

logger_default = get_logger("athos_runner")

ProgressCallback = Callable[[float, str], None]


@dataclass
class AthosRunResult:
    generated_files: List[Path]
    report_path: Optional[Path] = None


class AthosRunner:
    RULE_ORDER = [
        "FORA_DE_LINHA",
        "ESTOQUE_COMPARTILHADO",
        "ENVIO_IMEDIATO",
        "SEM_GRUPO",
        "OUTLET",
    ]

    OUTPUT_NAMES = {
        "FORA_DE_LINHA": "01_FORA_DE_LINHA.xlsx",
        "ESTOQUE_COMPARTILHADO": "02_ESTOQUE_COMPARTILHADO.xlsx",
        "ENVIO_IMEDIATO": "03_ENVIO_IMEDIATO.xlsx",
        "SEM_GRUPO": "04_SEM_GRUPO.xlsx",
        "OUTLET": "05_OUTLET.xlsx",
        "RELATORIO": "RELATORIO_CONSOLIDADO.xlsx",
    }

    def __init__(self, output_dir: Path, logger=logger_default):
        self.output_dir = Path(output_dir)
        self.logger = logger

    def run(
        self,
        sql_export_path: Path,
        whitelist_path: Path,
        template_path: Path,
        progress_callback: Optional[ProgressCallback] = None,
        send_email: bool = True,
    ) -> AthosRunResult:
        progress = progress_callback or (lambda p, m="": None)

        sql_export_path = Path(sql_export_path)
        whitelist_path = Path(whitelist_path)
        template_path = Path(template_path)

        progress(0.02, "Validando arquivos...")
        self._validate_inputs(sql_export_path, whitelist_path, template_path)

        self.output_dir.mkdir(parents=True, exist_ok=True)

        progress(0.08, "Lendo export do SQL...")
        sql_rows = self._read_excel_any(sql_export_path)

        progress(0.14, "Lendo whitelist...")
        from .athos_whitelist import load_whitelist
        wl = load_whitelist(whitelist_path)
        whitelist_eans = wl.eans

        self.logger.info(f"[AthosRunner] SQL rows: {len(sql_rows)}")
        self.logger.info(f"[AthosRunner] Whitelist EANs: {len(whitelist_eans)}")

        progress(0.22, "Aplicando regras...")
        from .athos_rules_engine import process_rows

        # ✅ Instancia DB 1x só (e não quebra se faltar)
        supplier_db = None
        try:
            from ..core.supplier_database import SupplierDatabase
            supplier_db = SupplierDatabase()
        except Exception as e:
            supplier_db = None
            self.logger.warning(f"[AthosRunner] SupplierDatabase indisponível: {e}")

        def prazo_lookup(marca: str) -> Optional[int]:
            if not marca:
                return None
            if supplier_db is None:
                return None
            try:
                s = supplier_db.search_supplier_by_name(marca)
                prazo = getattr(s, "prazo_dias", None) if s else None
                return int(prazo) if prazo is not None else None
            except Exception:
                return None

        outputs = process_rows(
            sql_rows=sql_rows,
            whitelist_imediatos=whitelist_eans,
            supplier_prazo_lookup=prazo_lookup,
        )

        progress(0.35, "Gerando planilhas...")
        from .athos_excel_writer import AthosExcelWriter
        from .athos_models import ORDERED_RULES, RuleName

        writer = AthosExcelWriter()
        generated_files: List[Path] = []

        # ✅ converter ações -> linhas do template (headers reais do seu modelo)
        def action_to_row(a) -> Dict[str, Any]:
            row: Dict[str, Any] = {
                "Código de Barras": getattr(a, "codbarra", None),
            }

            # ✅ NÃO preencher Tipo Produto (pedido)

            if getattr(a, "grupo3", None) is not None:
                row["GRUPO3"] = a.grupo3

            if getattr(a, "estoque_seguranca", None) is not None:
                row["Estoque de Segurança"] = a.estoque_seguranca

            if getattr(a, "produto_inativo", None) is not None:
                row["Produto Inativo"] = a.produto_inativo

            if getattr(a, "dias_entrega", None) is not None:
                row["Dias para Entrega"] = a.dias_entrega

            # ✅ REGRA: se dias_entrega == 0 -> Site Disponibilidade = "Imediata"
            site_disp = getattr(a, "site_disponibilidade", None)
            dias = getattr(a, "dias_entrega", None)

            if site_disp is not None:
                site = str(site_disp).strip()
                if (dias == 0) or (site.lower() == "imediata") or (site.upper() == "IMEDIATA"):
                    row["Site Disponibilidade"] = "Imediata"
                else:
                    row["Site Disponibilidade"] = site

            if dias == 0 and "Site Disponibilidade" not in row:
                row["Site Disponibilidade"] = "Imediata"

            return row

        # ✅ NORMALIZAÇÃO CRÍTICA:
        # outputs.actions_by_rule[rule] pode ser list[AthosAction] OU dict[codbarra, AthosAction]
        def _iter_actions(bucket: Any) -> Iterable[Any]:
            if bucket is None:
                return []
            if isinstance(bucket, dict):
                return list(bucket.values())
            if isinstance(bucket, list):
                return bucket
            # fallback iterável
            try:
                return list(bucket)
            except Exception:
                return []

        rule_to_output = {
            RuleName.FORA_DE_LINHA: "FORA_DE_LINHA",
            RuleName.ESTOQUE_COMPARTILHADO: "ESTOQUE_COMPARTILHADO",
            RuleName.ENVIO_IMEDIATO: "ENVIO_IMEDIATO",
            RuleName.NENHUM_GRUPO: "SEM_GRUPO",
            RuleName.OUTLET: "OUTLET",
        }

        actions_by_rule = getattr(outputs, "actions_by_rule", {}) or {}

        for idx, rule in enumerate(ORDERED_RULES, start=1):
            pct = 0.35 + (idx / len(ORDERED_RULES)) * 0.45
            progress(pct, f"Escrevendo {rule.value}...")

            key = rule_to_output[rule]
            out_path = self.output_dir / self.OUTPUT_NAMES[key]
            self._copy_template(template_path, out_path)

            bucket = actions_by_rule.get(rule, None)
            actions = _iter_actions(bucket)

            # log útil pra validar rapidamente se OUTLET está emitindo KIT/PAI
            try:
                tipos = []
                for a in actions:
                    t = getattr(a, "tipo", "")
                    tipos.append(t.value if hasattr(t, "value") else str(t))
                self.logger.info(f"[AthosRunner] {rule.value}: {len(actions)} actions | tipos={dict(Counter(tipos))}")
            except Exception:
                pass

            rows_to_write = []
            for a in actions:
                r = action_to_row(a)
                # descarta lixo (ex.: caso bucket tivesse string/chave)
                if not r.get("Código de Barras"):
                    continue
                rows_to_write.append(r)

            writer.write_rule_workbook(
                out_path,
                rows_to_write,
                sheet_name="PRODUTOS",
                clear_existing_data=True,
            )
            generated_files.append(out_path)

        progress(0.86, "Gerando relatório consolidado...")
        report_path = self.output_dir / self.OUTPUT_NAMES["RELATORIO"]
        self._write_report(report_path, outputs, sql_rows, sql_export_path, whitelist_path, template_path)

        if send_email:
            progress(0.92, "Enviando e-mail (RPA Athus)...")
            try:
                attachments = list(generated_files)
                if report_path and report_path.exists():
                    attachments.append(report_path)
                self._send_email_athos(attachments)
            except Exception as e:
                self.logger.warning(f"⚠️ Falha ao enviar e-mail do Athos: {e}")

        progress(1.0, "Concluído ✅")
        return AthosRunResult(generated_files=generated_files, report_path=report_path)

    # ===== IO =====
    def _validate_inputs(self, sql_export_path: Path, whitelist_path: Path, template_path: Path) -> None:
        if not sql_export_path.exists():
            raise FileNotFoundError(f"Arquivo do SQL não encontrado: {sql_export_path}")
        if not whitelist_path.exists():
            raise FileNotFoundError(f"Whitelist não encontrada: {whitelist_path}")
        if not template_path.exists():
            raise FileNotFoundError(f"Template não encontrado: {template_path}")
        if template_path.suffix.lower() not in [".xlsx", ".xlsm"]:
            raise ValueError("Template precisa ser .xlsx ou .xlsm (modelo Athos).")

    def _copy_template(self, template_path: Path, output_path: Path) -> None:
        shutil.copy2(template_path, output_path)

    def _read_excel_any(self, path: Path) -> List[Dict[str, Any]]:
        suffix = path.suffix.lower()
        if suffix in [".xlsx", ".xlsm"]:
            return self._read_xlsx_openpyxl(path)
        if suffix == ".xls":
            return self._read_xls_pandas(path)
        raise ValueError(f"Formato não suportado: {suffix}. Use .xlsx/.xlsm (ou .xls com suporte xlrd).")

    def _read_xlsx_openpyxl(self, path: Path) -> List[Dict[str, Any]]:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        headers: List[str] = []
        rows_out: List[Dict[str, Any]] = []

        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx == 1:
                headers = [str(c).strip() if c is not None else "" for c in row]
                continue

            if not any(c is not None and str(c).strip() != "" for c in row):
                continue

            item: Dict[str, Any] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                item[h] = row[i] if i < len(row) else None

            rows_out.append(item)

        return rows_out

    def _read_xls_pandas(self, path: Path) -> List[Dict[str, Any]]:
        try:
            import pandas as pd  # type: ignore
        except Exception:
            raise RuntimeError(
                "Não consegui ler arquivo .xls (pandas não disponível). "
                "Salve a whitelist como .xlsx e tente novamente."
            )

        try:
            df = pd.read_excel(path, engine="xlrd")  # type: ignore
        except Exception:
            try:
                df = pd.read_excel(path)  # type: ignore
            except Exception as e:
                raise RuntimeError(
                    "Não consegui ler arquivo .xls. "
                    "Sugestão: abra e salve como .xlsx. "
                    f"Detalhe: {e}"
                )

        df = df.dropna(how="all")
        return df.to_dict(orient="records")

    # ===== E-mail Athos =====
    def _send_email_athos(self, attachments: List[Path]) -> None:
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        import smtplib
        import ssl

        from ..core.config import load_config

        cfg = load_config()
        if not cfg.email:
            raise RuntimeError("Config de e-mail não encontrada em assets/config/settings.json")

        to_addr = "rpa.athus@apoiocorp.com.br"

        msg = MIMEMultipart()
        msg["From"] = cfg.email.from_addr or cfg.email.username
        msg["To"] = to_addr
        msg["Subject"] = "DROSSI PRODUTOS"
        msg.attach(MIMEText("Planilhas geradas pelo Robô Athos em anexo.", "plain", "utf-8"))

        for fp in attachments:
            if not fp.exists():
                continue
            part = MIMEBase("application", "octet-stream")
            part.set_payload(fp.read_bytes())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={fp.name}")
            msg.attach(part)

        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(cfg.email.smtp_host, cfg.email.smtp_port, context=context) as server:
            server.login(cfg.email.username, cfg.email.password)
            server.sendmail(msg["From"], [to_addr], msg.as_string())

    def _write_report(
        self,
        report_path: Path,
        outputs: Any,
        sql_rows: List[Dict[str, Any]],
        sql_export_path: Path,
        whitelist_path: Path,
        template_path: Path,
    ) -> None:
        """Wrapper: chama a implementação do relatório XLSX definida no módulo."""
        return globals()["_write_report"](
            self, report_path, outputs, sql_rows, sql_export_path, whitelist_path, template_path
        )


# ===== Report =====
def _write_report(
    self,
    report_path: Path,
    outputs: Any,
    sql_rows: List[Dict[str, Any]],
    sql_export_path: Path,
    whitelist_path: Path,
    template_path: Path,
) -> None:
    """Gera RELATORIO_CONSOLIDADO.xlsx (substitui o antigo .txt).

    Requisito: trazer CÓD AUXILIAR (PA/KIT/PAI) + Dias para Entrega e Site Disponibilidade.

    ✅ Correção: outputs.actions_by_rule pode vir como:
       - Dict[RuleName, List[AthosAction]]  (padrão do seu athos_rules_engine)
       - ou Dict[RuleName, Dict[codbarra, AthosAction]] (caso alguém altere no futuro)
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from .athos_models import normalize_ean

    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Map COD_BARRA -> COD_AUXILIAR (por tipo)
    codaux_pa: Dict[str, Any] = {}
    codaux_kit: Dict[str, Any] = {}
    codaux_pai: Dict[str, Any] = {}

    for row in sql_rows:
        pa = normalize_ean(row.get("CODBARRA_PRODUTO"))
        kit = normalize_ean(row.get("CODBARRA_KIT"))
        pai = normalize_ean(row.get("CODBARRA_PAI"))

        if pa and pa not in codaux_pa:
            codaux_pa[pa] = row.get("CODAUXILIAR_PRODUTO")
        if kit and kit not in codaux_kit:
            codaux_kit[kit] = row.get("CODAUXILIAR_KIT")
        if pai and pai not in codaux_pai:
            codaux_pai[pai] = row.get("CODAUXILIAR_PAI")

    def _cod_aux(tipo: str, codbarra: str) -> Any:
        if tipo == "PA":
            return codaux_pa.get(codbarra)
        if tipo == "KIT":
            return codaux_kit.get(codbarra)
        return codaux_pai.get(codbarra)

    def _iter_bucket(bucket: Any):
        """Normaliza o bucket para iterar como (codbarra, action)."""
        if bucket is None:
            return
        if isinstance(bucket, dict):
            for cb, a in bucket.items():
                yield cb, a
            return
        if isinstance(bucket, list):
            for a in bucket:
                cb = getattr(a, "codbarra", None) or ""
                yield cb, a
            return
        # fallback: tenta iterar
        try:
            for a in bucket:
                cb = getattr(a, "codbarra", None) or ""
                yield cb, a
        except Exception:
            return

    wb = Workbook()

    # Aba 1: METADADOS
    ws_meta = wb.active
    ws_meta.title = "META"
    meta_lines = [
        ("RELATÓRIO CONSOLIDADO — ROBÔ ATHOS", ""),
        ("Gerado em", now),
        ("SQL export", str(sql_export_path)),
        ("Whitelist", str(whitelist_path)),
        ("Template", str(template_path)),
    ]
    for i, (k, v) in enumerate(meta_lines, start=1):
        ws_meta.cell(row=i, column=1).value = k
        ws_meta.cell(row=i, column=2).value = v
    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 90

    # Aba 2: RELATORIO
    ws = wb.create_sheet("RELATORIO")
    headers = [
        "REGRA",
        "TIPO",
        "COD_BARRA",
        "COD_AUXILIAR",
        "MARCA",
        "GRUPO3_ORIGEM",
        "GRUPO3_DESTINO",
        "ESTOQUE_SEG",
        "PRODUTO_INATIVO",
        "DIAS_PARA_ENTREGA",
        "SITE_DISPONIBILIDADE",
        "ACAO",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = h

    r = 2
    actions_by_rule = getattr(outputs, "actions_by_rule", {}) or {}

    for rule, bucket in actions_by_rule.items():
        regra = rule.value if hasattr(rule, "value") else str(rule)

        for codbarra, a in _iter_bucket(bucket):
            if not codbarra:
                continue

            tipo_obj = getattr(a, "tipo", "")
            tipo = tipo_obj.value if hasattr(tipo_obj, "value") else str(tipo_obj)

            dias = getattr(a, "dias_entrega", None)
            site = getattr(a, "site_disponibilidade", None)

            # regra global: dias==0 => "Imediata"
            if dias == 0:
                site = "Imediata"
            if isinstance(site, str) and site.strip().upper() == "IMEDIATA":
                site = "Imediata"

            ws.cell(row=r, column=1).value = regra
            ws.cell(row=r, column=2).value = tipo
            ws.cell(row=r, column=3).value = codbarra
            ws.cell(row=r, column=4).value = _cod_aux(tipo, codbarra)
            ws.cell(row=r, column=5).value = getattr(a, "marca", "") or ""
            ws.cell(row=r, column=6).value = getattr(a, "grupo3_origem_pa", "") or ""
            ws.cell(row=r, column=7).value = getattr(a, "grupo3", "") or ""
            ws.cell(row=r, column=8).value = getattr(a, "estoque_seguranca", None)
            ws.cell(row=r, column=9).value = getattr(a, "produto_inativo", None)
            ws.cell(row=r, column=10).value = dias
            ws.cell(row=r, column=11).value = site

            msgs = getattr(a, "mensagens", None)
            ws.cell(row=r, column=12).value = "; ".join(msgs) if isinstance(msgs, list) else (msgs or "")

            r += 1

    ws.freeze_panes = "A2"

    widths = [22, 8, 20, 18, 22, 20, 20, 12, 14, 16, 18, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(report_path)