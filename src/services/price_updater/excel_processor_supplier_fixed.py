import pandas as pd
import os
import math
import re
from datetime import datetime
from openpyxl import load_workbook


class ExcelProcessorUnified:
    def __init__(self, mode="Fábrica"):
        self.mode = mode
        self.base_data = {}  # Estrutura: {codigo: {tc: {A: dados, B: dados, C: dados, D: dados...}}}

        # ✅ LINHAS TC SERÃO AUTO-DETECTADAS (APENAS LETRAS)
        self.supported_tc_lines = []
        self.detected_tc_stats = {}

        self.header_rows = {
            'Poltrona': 57,
            'Namoradeira-Sofá': 24,
            'Puff-Banqueta': 40,
            'Cadeira': 25
        }

    def clean_currency_value(self, value):
        """Limpar e converter valores monetários"""
        try:
            if pd.isna(value) or value == '' or value is None:
                return 0.0

            str_value = str(value).strip()

            if str_value == '' or str_value.lower() == 'nan':
                return 0.0

            cleaned = re.sub(r'[R$\s]', '', str_value)
            cleaned = cleaned.replace(',', '.')
            cleaned = re.sub(r'[^\d.]', '', cleaned)

            if cleaned == '' or cleaned == '.':
                return 0.0

            return float(cleaned)

        except (ValueError, TypeError) as e:
            return 0.0

    def check_network_path(self, path, log_callback=None):
        """Verificar acesso ao caminho de rede"""
        try:
            if log_callback:
                log_callback(f"🔍 Verificando acesso: {path}")

            if os.path.exists(path):
                if log_callback:
                    log_callback("✅ Caminho acessível")
                return True
            else:
                if log_callback:
                    log_callback("❌ Caminho não acessível")
                return False

        except Exception as e:
            if log_callback:
                log_callback(f"❌ Erro ao verificar rede: {e}")
            return False

    def auto_detect_tc_lines(self, base_file, log_callback=None):
        """AUTO-DETECTAR todas as linhas TC existentes na planilha (APENAS LETRAS)"""
        try:
            if log_callback:
                log_callback("🔍 AUTO-DETECTANDO linhas TC disponíveis...")

            excel_file = pd.ExcelFile(base_file)
            detected_tc_lines = set()
            tc_frequency = {}

            for sheet_name in excel_file.sheet_names:
                # Determinar linha do cabeçalho
                if self.mode == "Fábrica" and sheet_name in self.header_rows:
                    header_row = self.header_rows[sheet_name] - 1
                else:
                    header_row = 1

                try:
                    df = pd.read_excel(base_file, sheet_name=sheet_name, header=header_row)

                    if df.empty or 'TC' not in df.columns:
                        continue

                    # Analisar coluna TC
                    for index, row in df.iterrows():
                        tc = str(row['TC']).strip().upper() if pd.notna(row['TC']) else None

                        # ✅ ACEITAR APENAS LETRAS ÚNICAS (A-Z)
                        if tc and len(tc) == 1 and tc.isalpha():
                            detected_tc_lines.add(tc)
                            tc_frequency[tc] = tc_frequency.get(tc, 0) + 1

                except Exception as e:
                    if log_callback:
                        log_callback(f"⚠️ Erro ao analisar aba {sheet_name}: {e}")
                    continue

            # ✅ ORDENAR LINHAS TC DETECTADAS
            self.supported_tc_lines = sorted(list(detected_tc_lines))
            self.detected_tc_stats = tc_frequency

            excel_file.close()

            if log_callback:
                if self.supported_tc_lines:
                    log_callback(f"✅ LINHAS TC DETECTADAS: {', '.join(self.supported_tc_lines)}")
                    log_callback(f"📊 FREQUÊNCIAS: {dict(sorted(tc_frequency.items()))}")
                else:
                    log_callback("❌ NENHUMA LINHA TC VÁLIDA DETECTADA!")

            return len(self.supported_tc_lines) > 0

        except Exception as e:
            if log_callback:
                log_callback(f"❌ Erro na auto-detecção TC: {e}")
            return False

    def load_base_data(self, base_file, log_callback=None):
        """Carregar dados da planilha base COM AUTO-DETECÇÃO TC"""
        try:
            if log_callback:
                log_callback(f"📊 Carregando dados base ({self.mode}) COM AUTO-DETECÇÃO TC...")

            # ✅ PRIMEIRO: AUTO-DETECTAR LINHAS TC
            if not self.auto_detect_tc_lines(base_file, log_callback):
                raise Exception("❌ ERRO CRÍTICO: Não foi possível detectar nenhuma linha TC válida na planilha base!")

            excel_file = pd.ExcelFile(base_file)

            for sheet_name in excel_file.sheet_names:
                if log_callback:
                    log_callback(f"�� Processando aba: {sheet_name}")

                # Determinar linha do cabeçalho
                if self.mode == "Fábrica" and sheet_name in self.header_rows:
                    header_row = self.header_rows[sheet_name] - 1
                    if log_callback:
                        log_callback(f"🏭 Fábrica - Usando linha {self.header_rows[sheet_name]} como cabeçalho")
                else:
                    header_row = 1
                    if log_callback:
                        log_callback(f"�� Fornecedor - Usando linha 2 como cabeçalho")

                df = pd.read_excel(base_file, sheet_name=sheet_name, header=header_row)

                if df.empty:
                    if log_callback:
                        log_callback(f"⚠️ Aba {sheet_name} vazia")
                    continue

                # Verificar colunas obrigatórias
                required_cols = ['TC', 'Código Fabricante', 'Custo For', 'Custo Fre', 'Preço De']
                missing_columns = []
                for col in required_cols:
                    if col not in df.columns:
                        missing_columns.append(col)

                if missing_columns:
                    if log_callback:
                        log_callback(f"⚠️ Aba {sheet_name}: Colunas obrigatórias faltando: {missing_columns}")
                    continue

                # Verificar colunas opcionais
                has_ipi = 'IPI' in df.columns
                has_preco_por = 'Preço Por' in df.columns

                if self.mode == "Fornecedor" and has_ipi:
                    if log_callback:
                        log_callback(f"💰 Aba {sheet_name}: Coluna IPI encontrada")

                if has_preco_por:
                    if log_callback:
                        log_callback(f"📝 Aba {sheet_name}: Coluna 'Preço Por' encontrada")

                # ✅ PROCESSAR DADOS COM TC EXPANDIDO
                processed_count = 0
                tc_counts = {tc: 0 for tc in self.supported_tc_lines}
                tc_counts['outros'] = 0

                for index, row in df.iterrows():
                    codigo = str(row['Código Fabricante']).strip() if pd.notna(row['Código Fabricante']) else None
                    tc = str(row['TC']).strip().upper() if pd.notna(row['TC']) else None

                    # ✅ ACEITAR APENAS TC DETECTADOS
                    if codigo and codigo != 'nan' and tc and tc in self.supported_tc_lines:
                        # Inicializar estrutura do produto se não existir
                        if codigo not in self.base_data:
                            self.base_data[codigo] = {}

                        # Armazenar dados por TC (linha do tecido)
                        self.base_data[codigo][tc] = {
                            'custo_for': self.clean_currency_value(row['Custo For']),
                            'custo_fre': self.clean_currency_value(row['Custo Fre']),
                            'preco_de': self.clean_currency_value(row['Preço De']),
                            'preco_por': self.clean_currency_value(row['Preço Por']) if has_preco_por else 0.0,
                            'ipi': self.clean_currency_value(row['IPI']) if has_ipi else 0.0,
                            'aba': sheet_name,
                            'linha_planilha': index + header_row + 2,
                            'tc': tc
                        }

                        processed_count += 1
                        tc_counts[tc] += 1

                        # Log detalhado para primeiros 2 itens de cada TC
                        if tc_counts[tc] <= 2 and log_callback:
                            data = self.base_data[codigo][tc]
                            log_msg = f"💎 {codigo}({tc}): For R$ {data['custo_for']:.2f} + Fre R$ {data['custo_fre']:.2f}"
                            if data['ipi'] > 0:
                                log_msg += f" + IPI R$ {data['ipi']:.2f}"
                            log_msg += f" (De R$ {data['preco_de']:.2f}"
                            if data['preco_por'] > 0:
                                log_msg += f", Por R$ {data['preco_por']:.2f})"
                            else:
                                log_msg += ")"
                            log_callback(log_msg)
                    elif tc and tc not in self.supported_tc_lines:
                        tc_counts['outros'] += 1

                if log_callback:
                    log_callback(f"✅ Aba {sheet_name}: {processed_count} produtos processados")
                    # ✅ LOG EXPANDIDO PARA TODAS AS LINHAS TC
                    tc_summary = ", ".join([f"TC {tc}: {count}" for tc, count in tc_counts.items() if count > 0])
                    log_callback(f"📊 {tc_summary}")

            total_products = len(self.base_data)
            total_variants = sum(len(variants) for variants in self.base_data.values())

            if log_callback:
                log_callback(
                    f"🎉 TOTAL CARREGADO: {total_products} produtos únicos, {total_variants} variantes TC ({self.mode})")
                log_callback(f"🔤 LINHAS TC ATIVAS: {', '.join(self.supported_tc_lines)}")

            excel_file.close()
            return True

        except Exception as e:
            if log_callback:
                log_callback(f"❌ Erro ao carregar dados base: {e}")
            return False

    def apply_90_cents_rule(self, price):
        """Aplicar regra dos ,90 centavos para preços"""
        if pd.isna(price) or price <= 0:
            return 0.0

        integer_part = math.floor(float(price))
        final_price = integer_part + 0.90
        return final_price

    def extract_fabric_line_and_code(self, full_code):
        """Extrair linha do tecido (TC) e código base - SEM PADRÃO!"""
        if not full_code:
            return None, None

        full_code = str(full_code).strip()

        # ✅ VERIFICAR SE ÚLTIMA POSIÇÃO É UMA LINHA TC VÁLIDA
        if len(full_code) > 1 and full_code[-1] in self.supported_tc_lines:
            fabric_line = full_code[-1]
            base_code = full_code[:-1]
            return base_code, fabric_line

        # ❌ NÃO USAR PADRÃO - RETORNAR ERRO
        return None, None

    def get_product_data_by_tc(self, code, tc, log_callback=None):
        """Obter dados de um produto específico por TC"""
        if code in self.base_data and tc in self.base_data[code]:
            return self.base_data[code][tc]
        return None

    def process_simple_code(self, code, fabric_line, log_callback=None):
        """Processar código simples COM BUSCA POR TC OBRIGATÓRIA"""
        # ✅ VERIFICAR SE TC É VÁLIDO
        if not fabric_line or fabric_line not in self.supported_tc_lines:
            return {
                'found': False,
                'detail': f"TC inválido ou não encontrado: '{fabric_line}'. TCs disponíveis: {', '.join(self.supported_tc_lines)}"
            }

        product_data = self.get_product_data_by_tc(code, fabric_line, log_callback)

        if product_data:
            # Calcular os valores finais para o produto na linha específica
            vr_custo_total = product_data['custo_for']
            custo_frete = product_data['custo_fre']
            custo_ipi = product_data['ipi']
            preco_de_venda = product_data['preco_de']
            preco_promocao = product_data['preco_por'] if product_data['preco_por'] > 0 else product_data['preco_de']

            return {
                'vr_custo_total': vr_custo_total,
                'custo_frete': custo_frete,
                'custo_ipi': custo_ipi,
                'preco_de_venda': preco_de_venda,
                'preco_promocao': preco_promocao,
                'found': True,
                'detail': f"Encontrado {code}(TC {fabric_line}): For R$ {vr_custo_total:.2f}, Fre R$ {custo_frete:.2f}" + (
                    f", IPI R$ {custo_ipi:.2f}" if custo_ipi > 0 else "")
            }

        return {
            'found': False,
            'detail': f"Código não encontrado: {code} na linha TC {fabric_line}"
        }

    def process_multiplied_code(self, multiplier, code, fabric_line, log_callback=None):
        """Processar código com multiplicador COM TC OBRIGATÓRIO"""
        simple_result = self.process_simple_code(code, fabric_line, log_callback)

        if simple_result['found']:
            return {
                'vr_custo_total': simple_result['vr_custo_total'] * multiplier,
                'custo_frete': simple_result['custo_frete'] * multiplier,
                'custo_ipi': simple_result['custo_ipi'] * multiplier,
                'preco_de_venda': simple_result['preco_de_venda'] * multiplier,
                'preco_promocao': simple_result['preco_promocao'] * multiplier,
                'found': True,
                'detail': f"Multiplicado {multiplier}x: {simple_result['detail']}"
            }

        return simple_result

    def process_kit_with_bars(self, kit_str, log_callback=None):
        """Processar kit com barras COM TC OBRIGATÓRIO"""
        base_kit, fabric_line = self.extract_fabric_line_and_code(kit_str)

        # ✅ VERIFICAR SE TC FOI EXTRAÍDO CORRETAMENTE
        if not base_kit or not fabric_line:
            return {
                'found': False,
                'detail': f"TC não identificado no kit: {kit_str}. TCs disponíveis: {', '.join(self.supported_tc_lines)}"
            }

        components = base_kit.split('/')

        total_vr_custo_total = 0.0
        total_custo_frete = 0.0
        total_custo_ipi = 0.0
        total_preco_de_venda = 0.0
        total_preco_promocao = 0.0

        components_found = 0
        detail_parts = []

        for component in components:
            component = component.strip()
            if not component:
                continue

            multiplier = 1
            actual_code = component

            if '*' in component:
                parts = component.split('*')
                if len(parts) == 2:
                    try:
                        multiplier = float(parts[0])
                        actual_code = parts[1].strip()
                    except ValueError:
                        detail_parts.append(f"{component}: FORMATO MULTIPLICADOR INVÁLIDO")
                        continue

            # Usar o método simples para obter os valores base para cada componente COM TC
            simple_result = self.process_simple_code(actual_code, fabric_line, log_callback)

            if simple_result['found']:
                total_vr_custo_total += simple_result['vr_custo_total'] * multiplier
                total_custo_frete += simple_result['custo_frete'] * multiplier
                total_custo_ipi += simple_result['custo_ipi'] * multiplier
                total_preco_de_venda += simple_result['preco_de_venda'] * multiplier
                total_preco_promocao += simple_result['preco_promocao'] * multiplier
                components_found += 1
                detail_parts.append(f"{component}(TC {fabric_line}): OK")
            else:
                detail_parts.append(f"{component}(TC {fabric_line}): NÃO ENCONTRADO")

        if components_found > 0:
            return {
                'vr_custo_total': total_vr_custo_total,
                'custo_frete': total_custo_frete,
                'custo_ipi': total_custo_ipi,
                'preco_de_venda': total_preco_de_venda,
                'preco_promocao': total_preco_promocao,
                'found': True,
                'detail': f"Kit TC {fabric_line} processado ({components_found}/{len(components)} encontrados): {' | '.join(detail_parts)}"
            }

        return {
            'found': False,
            'detail': f"Nenhum componente encontrado no kit TC {fabric_line}: {kit_str}"
        }

    def process_code(self, code_str, log_callback=None):
        """Função unificada para processar qualquer tipo de código COM TC OBRIGATÓRIO"""
        if not code_str or pd.isna(code_str) or str(code_str).strip() == '':
            return {'found': False, 'detail': "Código vazio"}

        code_str = str(code_str).strip()

        # Verificar se é kit (tem barras)
        if '/' in code_str:
            return self.process_kit_with_bars(code_str, log_callback)

        # Verificar se tem multiplicador
        if '*' in code_str:
            parts = code_str.split('*')
            if len(parts) == 2:
                try:
                    multiplier = float(parts[0])
                    base_code_with_line = parts[1].strip()
                    base_code, fabric_line = self.extract_fabric_line_and_code(base_code_with_line)

                    # ✅ VERIFICAR SE TC FOI EXTRAÍDO
                    if not base_code or not fabric_line:
                        return {
                            'found': False,
                            'detail': f"TC não identificado em: {base_code_with_line}. TCs disponíveis: {', '.join(self.supported_tc_lines)}"
                        }

                    return self.process_multiplied_code(multiplier, base_code, fabric_line, log_callback)
                except ValueError:
                    return {'found': False, 'detail': f"Formato de multiplicador inválido: {code_str}"}

        # Código simples
        base_code, fabric_line = self.extract_fabric_line_and_code(code_str)

        # ✅ VERIFICAR SE TC FOI EXTRAÍDO
        if not base_code or not fabric_line:
            return {
                'found': False,
                'detail': f"TC não identificado em: {code_str}. TCs disponíveis: {', '.join(self.supported_tc_lines)}"
            }

        return self.process_simple_code(base_code, fabric_line, log_callback)

    def save_preserving_formatting_sequential(self, df_products, products_file, log_callback=None):
        """Método sequencial garantido - processa TODAS as linhas"""
        try:
            if log_callback:
                log_callback("💾 Salvando arquivo (método sequencial garantido)...")

            wb = load_workbook(products_file)
            ws = wb['PRODUTO']

            # Encontrar cabeçalho
            header_row_excel = None
            col_indices = {}

            for row_num in range(1, 6):
                found_codigo = False
                temp_indices = {}

                for col_idx, cell in enumerate(ws[row_num], 1):
                    if cell.value:
                        cell_value = str(cell.value).strip()
                        if cell_value == 'Código Fabricante':
                            found_codigo = True
                            temp_indices['Código Fabricante'] = col_idx
                        elif cell_value == 'VR Custo Total':
                            temp_indices['VR Custo Total'] = col_idx
                        elif cell_value == 'Custo Frete':
                            temp_indices['Custo Frete'] = col_idx
                        elif cell_value == 'Custo IPI':
                            temp_indices['Custo IPI'] = col_idx
                        elif cell_value == 'Preço de Venda':
                            temp_indices['Preço de Venda'] = col_idx
                        elif cell_value == 'Preço Promoção':
                            temp_indices['Preço Promoção'] = col_idx

                if found_codigo:
                    header_row_excel = row_num
                    col_indices = temp_indices
                    break

            if not header_row_excel:
                raise Exception("Cabeçalho não encontrado")

            if log_callback:
                log_callback(f"🔍 Cabeçalho encontrado na linha {header_row_excel}")
                log_callback(f"📋 Colunas mapeadas: {list(col_indices.keys())}")

            # MÉTODO SEQUENCIAL: Processar linha por linha do DataFrame
            updated_count = 0
            skipped_count = 0
            error_count = 0

            total_df_rows = len(df_products)

            for df_index, df_row in df_products.iterrows():
                try:
                    # Linha correspondente no Excel (assumindo mesma ordem)
                    excel_row = header_row_excel + 1 + df_index

                    # Verificar se não excedeu o máximo do Excel
                    if excel_row > ws.max_row:
                        if log_callback:
                            log_callback(f"⚠️ DataFrame tem mais linhas que Excel. Parando na linha {excel_row}")
                        break

                    codigo_df = str(df_row['Código Fabricante']).strip() if pd.notna(
                        df_row['Código Fabricante']) else ""

                    if not codigo_df:
                        skipped_count += 1
                        continue

                    # Verificar se tem dados para atualizar
                    has_data_to_update = False
                    for col in ['VR Custo Total', 'Custo Frete', 'Custo IPI', 'Preço de Venda', 'Preço Promoção']:
                        if col in df_products.columns and pd.notna(df_row[col]) and df_row[col] != 0:
                            has_data_to_update = True
                            break

                    if not has_data_to_update:
                        skipped_count += 1
                        if updated_count < 5 and log_callback:
                            log_callback(f"ℹ️ {codigo_df} (linha {excel_row}): Sem dados para atualizar")
                        continue

                    # ATUALIZAR TODAS AS COLUNAS NECESSÁRIAS
                    cells_updated = 0

                    # VR Custo Total
                    if 'VR Custo Total' in col_indices and 'VR Custo Total' in df_products.columns:
                        if pd.notna(df_row['VR Custo Total']) and df_row['VR Custo Total'] != 0:
                            cell = ws.cell(row=excel_row, column=col_indices['VR Custo Total'])
                            cell.value = float(df_row['VR Custo Total'])
                            cells_updated += 1

                    # Custo Frete
                    if 'Custo Frete' in col_indices and 'Custo Frete' in df_products.columns:
                        if pd.notna(df_row['Custo Frete']) and df_row['Custo Frete'] != 0:
                            cell = ws.cell(row=excel_row, column=col_indices['Custo Frete'])
                            cell.value = float(df_row['Custo Frete'])
                            cells_updated += 1

                    # Custo IPI
                    if 'Custo IPI' in col_indices and 'Custo IPI' in df_products.columns:
                        if pd.notna(df_row['Custo IPI']) and df_row['Custo IPI'] != 0:
                            cell = ws.cell(row=excel_row, column=col_indices['Custo IPI'])
                            cell.value = float(df_row['Custo IPI'])
                            cells_updated += 1

                    # Preço de Venda
                    if 'Preço de Venda' in col_indices and 'Preço de Venda' in df_products.columns:
                        if pd.notna(df_row['Preço de Venda']) and df_row['Preço de Venda'] != 0:
                            cell = ws.cell(row=excel_row, column=col_indices['Preço de Venda'])
                            cell.value = float(df_row['Preço de Venda'])
                            cells_updated += 1

                    # Preço Promoção
                    if 'Preço Promoção' in col_indices and 'Preço Promoção' in df_products.columns:
                        if pd.notna(df_row['Preço Promoção']) and df_row['Preço Promoção'] != 0:
                            cell = ws.cell(row=excel_row, column=col_indices['Preço Promoção'])
                            cell.value = float(df_row['Preço Promoção'])
                            cells_updated += 1

                    if cells_updated > 0:
                        updated_count += 1

                        # Log detalhado para primeiras 10 atualizações
                        if updated_count <= 10 and log_callback:
                            log_callback(
                                f"✅ {codigo_df} (linha Excel {excel_row}): {cells_updated} células atualizadas")

                            # Mostrar valores atualizados
                            if 'VR Custo Total' in df_products.columns and pd.notna(df_row['VR Custo Total']):
                                log_callback(f"   💰 VR Custo Total: R$ {df_row['VR Custo Total']:.2f}")
                            if 'Preço de Venda' in df_products.columns and pd.notna(df_row['Preço de Venda']):
                                log_callback(f"   🏷️ Preço de Venda: R$ {df_row['Preço de Venda']:.2f}")
                    else:
                        skipped_count += 1

                    # Log de progresso a cada 100 linhas
                    if (df_index + 1) % 100 == 0 and log_callback:
                        log_callback(f"🔄 Processadas {df_index + 1}/{total_df_rows} linhas...")

                except Exception as e:
                    error_count += 1
                    if error_count <= 5 and log_callback:
                        log_callback(f"❌ Erro na linha {df_index + 1}: {e}")
                    continue

            # Salvar arquivo
            wb.save(products_file)
            wb.close()

            if log_callback:
                log_callback(f"📊 RESULTADO FINAL:")
                log_callback(f"✅ Linhas atualizadas: {updated_count}")
                log_callback(f"⏭️ Linhas puladas: {skipped_count}")
                log_callback(f"❌ Erros: {error_count}")
                log_callback(f"📋 Total processado: {total_df_rows}")
                log_callback(f"💾 Arquivo salvo com formatação preservada!")

            return True

        except Exception as e:
            if log_callback:
                log_callback(f"❌ Erro no método sequencial: {e}")
            return False

    def process_files(self, products_file, base_file, progress_callback=None, log_callback=None):
        """Processar arquivos COM LÓGICA TC OBRIGATÓRIA E PRESERVAÇÃO DE FORMATAÇÃO"""
        try:
            if log_callback:
                log_callback(f"🚀 INICIANDO PROCESSAMENTO {self.mode.upper()} COM TC OBRIGATÓRIO...")

            if not self.check_network_path(base_file, log_callback):
                raise Exception("Não foi possível acessar a planilha base no servidor.")

            if not self.load_base_data(base_file, log_callback):
                raise Exception("Erro ao carregar dados da planilha base.")

            if progress_callback:
                progress_callback(20)

            if log_callback:
                log_callback("📊 Lendo aba 'PRODUTO' da planilha de produtos...")

            try:
                excel_file_products = pd.ExcelFile(products_file)
                available_sheets = excel_file_products.sheet_names
                excel_file_products.close()

                if 'PRODUTO' not in available_sheets:
                    if log_callback:
                        log_callback(f"❌ Aba 'PRODUTO' não encontrada. Abas disponíveis: {available_sheets}")
                    raise Exception(f"Aba 'PRODUTO' não encontrada na planilha de produtos.")

                df_products = pd.read_excel(products_file, sheet_name='PRODUTO')
                if log_callback:
                    log_callback("✅ Aba 'PRODUTO' carregada com sucesso!")

            except Exception as e:
                raise Exception(f"Erro ao ler planilha de produtos: {e}")

            if df_products.empty:
                raise Exception("Aba 'PRODUTO' está vazia!")

            if log_callback:
                log_callback(f"📋 Aba 'PRODUTO': {len(df_products)} linhas encontradas.")

            # Verificar colunas na planilha de produtos
            product_output_cols = {
                'Código Fabricante': 'str',
                'VR Custo Total': 'float',
                'Custo IPI': 'float',
                'Custo Frete': 'float',
                'Preço de Venda': 'float',
                'Preço Promoção': 'float'
            }

            missing_cols_products = []
            for col_name in product_output_cols.keys():
                if col_name not in df_products.columns:
                    missing_cols_products.append(col_name)

            if missing_cols_products:
                if log_callback:
                    log_callback(f"🔍 Colunas encontradas na aba 'PRODUTO': {list(df_products.columns)}")
                raise Exception(f"Colunas obrigatórias faltando na aba 'PRODUTO': {missing_cols_products}")

            if log_callback:
                log_callback("✅ Todas as colunas necessárias encontradas na aba 'PRODUTO'.")

            if progress_callback:
                progress_callback(40)

            updated_count = 0
            not_found_count = 0
            kits_processed = 0
            tc_invalid_count = 0  # ✅ CONTADOR DE TCs INVÁLIDOS
            # ✅ ESTATÍSTICAS TC DINÂMICAS
            tc_stats = {tc: 0 for tc in self.supported_tc_lines}

            total_rows = len(df_products)

            for index, row in df_products.iterrows():
                if (index + 1) % 100 == 0 and log_callback:
                    log_callback(f"🔄 Processando linha {index + 1} de {total_rows}...")

                codigo_produto = row['Código Fabricante']

                if pd.isna(codigo_produto) or str(codigo_produto).strip() == '':
                    continue

                result = self.process_code(codigo_produto, log_callback)

                if result['found']:
                    # Atribuir valores individualmente
                    df_products.at[index, 'VR Custo Total'] = result['vr_custo_total']
                    df_products.at[index, 'Custo Frete'] = result['custo_frete']
                    df_products.at[index, 'Custo IPI'] = result['custo_ipi']

                    # Aplicar regra dos ,90 apenas nos preços
                    df_products.at[index, 'Preço de Venda'] = self.apply_90_cents_rule(result['preco_de_venda'])
                    df_products.at[index, 'Preço Promoção'] = self.apply_90_cents_rule(result['preco_promocao'])

                    updated_count += 1

                    # ✅ ESTATÍSTICAS TC EXPANDIDAS
                    _, tc = self.extract_fabric_line_and_code(str(codigo_produto))
                    if tc and tc in tc_stats:
                        tc_stats[tc] += 1

                    if '/' in str(codigo_produto):
                        kits_processed += 1

                    if log_callback and updated_count <= 5:  # Log detalhado para primeiros 5
                        log_callback(
                            f"✅ {codigo_produto}: {result['detail']} -> VR Custo Total: R$ {result['vr_custo_total']:.2f}, Preço de Venda: R$ {self.apply_90_cents_rule(result['preco_de_venda']):.2f}")
                else:
                    not_found_count += 1

                    # ✅ VERIFICAR SE É ERRO DE TC INVÁLIDO
                    if "TC não identificado" in result['detail'] or "TC inválido" in result['detail']:
                        tc_invalid_count += 1

                    if log_callback and not_found_count <= 5:  # Log detalhado para primeiros 5 não encontrados
                        log_callback(f"❌ {codigo_produto}: {result['detail']}")

                if progress_callback:
                    progress = 40 + int((index / total_rows) * 50)
                    progress_callback(progress)

            if log_callback:
                # ✅ LOG EXPANDIDO PARA TODAS AS LINHAS TC PROCESSADAS
                tc_summary = ", ".join([f"{tc}: {count}" for tc, count in tc_stats.items() if count > 0])
                log_callback(f"📊 Estatísticas TC processadas: {tc_summary}")

                if tc_invalid_count > 0:
                    log_callback(f"⚠️ CÓDIGOS COM TC INVÁLIDO: {tc_invalid_count}")
                    log_callback(f"🔤 TCs VÁLIDOS: {', '.join(self.supported_tc_lines)}")

            if progress_callback:
                progress_callback(95)

            # USAR MÉTODO SEQUENCIAL GARANTIDO
            if not self.save_preserving_formatting_sequential(df_products, products_file, log_callback):
                if log_callback:
                    log_callback("⚠️ Método sequencial falhou. Usando pandas...")

                # Fallback final
                with pd.ExcelWriter(products_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_products.to_excel(writer, sheet_name='PRODUTO', index=False)

                if log_callback:
                    log_callback("✅ Salvamento alternativo concluído (formatação pode ter sido perdida)")

            if progress_callback:
                progress_callback(100)

            return {
                'updated': updated_count,
                'not_found': not_found_count,
                'kits_processed': kits_processed,
                'tc_invalid': tc_invalid_count,  # ✅ NOVO CAMPO
                'output_file': products_file
            }

        except Exception as e:
            if log_callback:
                log_callback(f"❌ Erro final no processamento: {e}")
            raise e